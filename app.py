
import os
import io
import uuid
import hashlib
import datetime as dt
import streamlit as st
import pandas as pd
import altair as alt  # Interactive charts
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter


# -------------------------------------
# Configuration & Constants
# -------------------------------------

# --- Storage root (env-aware) ---
# If you attach a Persistent Disk on Render, set DISK_PATH to its mount, e.g., /var/data
BASE_DIR = os.getenv("DISK_PATH", "data")



APP_NAME = "Scorecard Data Manager"

# Rebind all dataset paths to BASE_DIR
DATA_DIR = BASE_DIR
ATTACHMENTS_DIR = os.path.join(DATA_DIR, "attachments")
HISTORY_FILE = os.path.join(DATA_DIR, "history.xlsx")
COMBINED_FILE = os.path.join(DATA_DIR, "combined_data.xlsx")
COMBINED_FILE_CSV = os.path.join(DATA_DIR, "combined_data.csv")
AUDIT_LOG_FILE = os.path.join(DATA_DIR, "audit_log.xlsx")


# Export filename prefix for downloads (easy to change in one place)
EXPORT_PREFIX = "associates_"


# BA-specific
BA_ATTACHMENTS_DIR = os.path.join(DATA_DIR, "attachments_ba")
BA_HISTORY_FILE = os.path.join(DATA_DIR, "ba_history.xlsx")
BA_COMBINED_FILE = os.path.join(DATA_DIR, "ba_combined_data.xlsx")
BA_COMBINED_FILE_CSV = os.path.join(DATA_DIR, "ba_combined_data.csv")
BA_AUDIT_LOG_FILE = os.path.join(DATA_DIR, "ba_audit_log.xlsx")
BA_FEEDBACK_FILE = os.path.join(DATA_DIR, "ba_monthly_feedback.xlsx")
BA_EXPORT_PREFIX = "ba_"



# PE-specific
PE_ATTACHMENTS_DIR = os.path.join(DATA_DIR, "attachments_pe")
PE_HISTORY_FILE = os.path.join(DATA_DIR, "pe_history.xlsx")
PE_COMBINED_FILE = os.path.join(DATA_DIR, "pe_combined_data.xlsx")
PE_COMBINED_FILE_CSV = os.path.join(DATA_DIR, "pe_combined_data.csv")
PE_AUDIT_LOG_FILE = os.path.join(DATA_DIR, "pe_audit_log.xlsx")
PE_FEEDBACK_FILE = os.path.join(DATA_DIR, "pe_monthly_feedback.xlsx")
PE_EXPORT_PREFIX = "pe_"

# TL-specific
TL_ATTACHMENTS_DIR = os.path.join(DATA_DIR, "attachments_tl")
TL_HISTORY_FILE = os.path.join(DATA_DIR, "tl_history.xlsx")
TL_COMBINED_FILE = os.path.join(DATA_DIR, "tl_combined_data.xlsx")
TL_COMBINED_FILE_CSV = os.path.join(DATA_DIR, "tl_combined_data.csv")
TL_AUDIT_LOG_FILE = os.path.join(DATA_DIR, "tl_audit_log.xlsx")
TL_FEEDBACK_FILE = os.path.join(DATA_DIR, "tl_monthly_feedback.xlsx")
TL_EXPORT_PREFIX = "tl_"

# PL-specific
PL_ATTACHMENTS_DIR = os.path.join(DATA_DIR, "attachments_pl")
PL_HISTORY_FILE = os.path.join(DATA_DIR, "pl_history.xlsx")
PL_COMBINED_FILE = os.path.join(DATA_DIR, "pl_combined_data.xlsx")
PL_COMBINED_FILE_CSV = os.path.join(DATA_DIR, "pl_combined_data.csv")
PL_AUDIT_LOG_FILE = os.path.join(DATA_DIR, "pl_audit_log.xlsx")
PL_FEEDBACK_FILE = os.path.join(DATA_DIR, "pl_monthly_feedback.xlsx")
PL_EXPORT_PREFIX = "pl_"



# Excel limits (XLSX hard limits)
EXCEL_MAX_ROWS = 1048576
EXCEL_MAX_COLS = 16384
REQUIRED_COLS = ["Domain ID", "Function", "Function Lead", "Team Lead"]
MAX_UPLOAD_MB = 25
USERS = {
    "admin": {"password_hash": hashlib.sha256("admin123".encode()).hexdigest(), "role": "admin", "display_name": "Administrator"},
    "viewer": {"password_hash": hashlib.sha256("viewer123".encode()).hexdigest(), "role": "user", "display_name": "Viewer"},
}
# Feedback constants
FEEDBACK_FILE = os.path.join(DATA_DIR, "monthly_feedback.xlsx")
FEEDBACK_PASSWORD = "TL@2025"
MAX_FEEDBACK_CHARS = 500

# -------------------------------------
# Storage Setup
# -------------------------------------
def ensure_storage():
    os.makedirs(DATA_DIR, exist_ok=True)
    os.makedirs(ATTACHMENTS_DIR, exist_ok=True)
    if not os.path.exists(HISTORY_FILE):
        pd.DataFrame(columns=[
            "id","filename","saved_path","uploader","upload_dt","reporting_month",
            "rows_count","source_url","status","message","active","superseded_by","validation_status"
        ]).to_excel(HISTORY_FILE, index=False)
    if not os.path.exists(COMBINED_FILE):
        pd.DataFrame(columns=["Attachment ID"]).to_excel(COMBINED_FILE, index=False)
    if not os.path.exists(COMBINED_FILE_CSV):
        pd.DataFrame(columns=["Attachment ID"]).to_csv(COMBINED_FILE_CSV, index=False)
    if not os.path.exists(AUDIT_LOG_FILE):
        pd.DataFrame(columns=["timestamp","action","attachment_id","filename","performed_by"]).to_excel(AUDIT_LOG_FILE, index=False)
    if not os.path.exists(FEEDBACK_FILE):
        pd.DataFrame(columns=[
            "Domain ID","Name","Month","Team Lead","Feedback","timestamp","entered_by"
        ]).to_excel(FEEDBACK_FILE, index=False)

# BA storage initialization
def ensure_storage_ba():
    os.makedirs(BA_ATTACHMENTS_DIR, exist_ok=True)
    if not os.path.exists(BA_HISTORY_FILE):
        pd.DataFrame(columns=[
            "id","filename","saved_path","uploader","upload_dt","reporting_month",
            "rows_count","source_url","status","message","active","superseded_by","validation_status"
        ]).to_excel(BA_HISTORY_FILE, index=False)
    if not os.path.exists(BA_COMBINED_FILE):
        pd.DataFrame(columns=["Attachment ID"]).to_excel(BA_COMBINED_FILE, index=False)
    if not os.path.exists(BA_COMBINED_FILE_CSV):
        pd.DataFrame(columns=["Attachment ID"]).to_csv(BA_COMBINED_FILE_CSV, index=False)
    if not os.path.exists(BA_AUDIT_LOG_FILE):
        pd.DataFrame(columns=["timestamp","action","attachment_id","filename","performed_by"]).to_excel(BA_AUDIT_LOG_FILE, index=False)
    if not os.path.exists(BA_FEEDBACK_FILE):
        pd.DataFrame(columns=[
            "Domain ID","Name","Month","Team Lead","Feedback","timestamp","entered_by"
        ]).to_excel(BA_FEEDBACK_FILE, index=False)


# PE storage initialization
def ensure_storage_pe():
    os.makedirs(PE_ATTACHMENTS_DIR, exist_ok=True)
    if not os.path.exists(PE_HISTORY_FILE):
        pd.DataFrame(columns=[
            "id","filename","saved_path","uploader","upload_dt","reporting_month",
            "rows_count","source_url","status","message","active","superseded_by","validation_status"
        ]).to_excel(PE_HISTORY_FILE, index=False)
    if not os.path.exists(PE_COMBINED_FILE):
        pd.DataFrame(columns=["Attachment ID"]).to_excel(PE_COMBINED_FILE, index=False)
    if not os.path.exists(PE_COMBINED_FILE_CSV):
        pd.DataFrame(columns=["Attachment ID"]).to_csv(PE_COMBINED_FILE_CSV, index=False)
    if not os.path.exists(PE_AUDIT_LOG_FILE):
        pd.DataFrame(columns=["timestamp","action","attachment_id","filename","performed_by"]).to_excel(PE_AUDIT_LOG_FILE, index=False)
    if not os.path.exists(PE_FEEDBACK_FILE):
        pd.DataFrame(columns=[
            "Domain ID","Name","Month","Team Lead","Feedback","timestamp","entered_by"
        ]).to_excel(PE_FEEDBACK_FILE, index=False)


# TL storage initialization
def ensure_storage_tl():
    os.makedirs(TL_ATTACHMENTS_DIR, exist_ok=True)
    if not os.path.exists(TL_HISTORY_FILE):
        pd.DataFrame(columns=[
            "id","filename","saved_path","uploader","upload_dt","reporting_month",
            "rows_count","source_url","status","message","active","superseded_by","validation_status"
        ]).to_excel(TL_HISTORY_FILE, index=False)
    if not os.path.exists(TL_COMBINED_FILE):
        pd.DataFrame(columns=["Attachment ID"]).to_excel(TL_COMBINED_FILE, index=False)
    if not os.path.exists(TL_COMBINED_FILE_CSV):
        pd.DataFrame(columns=["Attachment ID"]).to_csv(TL_COMBINED_FILE_CSV, index=False)
    if not os.path.exists(TL_AUDIT_LOG_FILE):
        pd.DataFrame(columns=["timestamp","action","attachment_id","filename","performed_by"]).to_excel(TL_AUDIT_LOG_FILE, index=False)
    if not os.path.exists(TL_FEEDBACK_FILE):
        pd.DataFrame(columns=[
            "Domain ID","Name","Month","Team Lead","Feedback","timestamp","entered_by"
        ]).to_excel(TL_FEEDBACK_FILE, index=False)


# PL storage initialization
def ensure_storage_pl():
    os.makedirs(PL_ATTACHMENTS_DIR, exist_ok=True)
    if not os.path.exists(PL_HISTORY_FILE):
        pd.DataFrame(columns=[
            "id","filename","saved_path","uploader","upload_dt","reporting_month",
            "rows_count","source_url","status","message","active","superseded_by","validation_status"
        ]).to_excel(PL_HISTORY_FILE, index=False)
    if not os.path.exists(PL_COMBINED_FILE):
        pd.DataFrame(columns=["Attachment ID"]).to_excel(PL_COMBINED_FILE, index=False)
    if not os.path.exists(PL_COMBINED_FILE_CSV):
        pd.DataFrame(columns=["Attachment ID"]).to_csv(PL_COMBINED_FILE_CSV, index=False)
    if not os.path.exists(PL_AUDIT_LOG_FILE):
        pd.DataFrame(columns=["timestamp","action","attachment_id","filename","performed_by"]).to_excel(PL_AUDIT_LOG_FILE, index=False)
    if not os.path.exists(PL_FEEDBACK_FILE):
        pd.DataFrame(columns=[
            "Domain ID","Name","Month","Team Lead","Feedback","timestamp","entered_by"
        ]).to_excel(PL_FEEDBACK_FILE, index=False)


# -------------------------------------
# Authentication
# -------------------------------------
def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

def authenticate(username, password):
    return (True, USERS[username]["role"]) if username in USERS and hash_password(password) == USERS[username]["password_hash"] else (False, None)

# -------------------------------------
# Excel Helpers
# -------------------------------------
def safe_month_from_columns(df):
    for col in ["Month","Date"]:
        if col in df.columns:
            s = pd.to_datetime(df[col], errors="coerce")
            if s.notna().any():
                d = s.dropna().iloc[0]
                return f"{d.year:04d}-{d.month:02d}"
    now = dt.datetime.now()
    return f"{now.year:04d}-{now.month:02d}"

def detect_header_and_read(xls, sheet_name):
    raw = pd.read_excel(xls, sheet_name=sheet_name, header=None)
    header_idx = 0
    for i in range(min(10, len(raw))):
        row = [str(v).strip() for v in raw.iloc[i].tolist()]
        if all(col in row for col in REQUIRED_COLS):
            header_idx = i
            break
    df = pd.read_excel(xls, sheet_name=sheet_name, header=header_idx)
    df.columns = df.columns.str.strip()
    return df

def read_excel_bytes(file_bytes):
    try:
        xls = pd.ExcelFile(io.BytesIO(file_bytes), engine="openpyxl")
    except Exception as e:
        raise ValueError(f"Invalid Excel file: {e}")
    if "Data" not in xls.sheet_names:
        raise ValueError("Missing required sheet: Data")
    return detect_header_and_read(xls, "Data")

def validate_required_columns(df):
    return [col for col in REQUIRED_COLS if col.lower() not in [c.lower() for c in df.columns]]

def load_history():
    return pd.read_excel(HISTORY_FILE)

def save_history(df):
    df.to_excel(HISTORY_FILE, index=False)

# Robust coercion for the History.active column (handles True/False, 1/0, "TRUE"/"FALSE", "yes"/"no")
def _coerce_active_bool(series):
    if series is None:
        return pd.Series(False, index=[])
    s = series.copy()
    s = s.astype(str).str.strip().str.lower()
    truthy = {"true","1","yes","y","t"}
    falsy = {"false","0","no","n","f",""}
    return s.apply(lambda v: True if v in truthy else (False if v in falsy else False))

# Prefer XLSX; if missing or failed, fall back to CSV
def load_combined():
    try:
        return pd.read_excel(COMBINED_FILE)
    except Exception:
        try:
            return pd.read_csv(COMBINED_FILE_CSV)
        except Exception:
            return pd.DataFrame(columns=["Attachment ID"])

# Save to XLSX if within Excel bounds; otherwise save to CSV to avoid hard Excel limits
def save_combined(df):
    if not exceeds_excel_limits(df):
        df.to_excel(COMBINED_FILE, index=False)
    else:
        df.to_csv(COMBINED_FILE_CSV, index=False)

def derive_saved_path(month, name):
    return os.path.join(ATTACHMENTS_DIR, f"{month}_{name.replace('/', '_').replace(chr(92), '_')}")

def supersede_existing_month(month, new_id, history_df):
    mask = (history_df["reporting_month"] == month) & (history_df["active"] == True)
    for idx in history_df[mask].index:
        try: os.remove(history_df.at[idx, "saved_path"])
        except: pass
        history_df.at[idx, "active"] = False
        history_df.at[idx, "superseded_by"] = new_id
    return history_df

# -------------------------------------
# Percentage Conversion & Numeric Companions
# -------------------------------------
def convert_percentage_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    def looks_like_percent_col(name: str) -> bool:
        n = name.strip().lower()
        return ("target" in n or "actual" in n or "rating" in n or "final score" in n
                or n.endswith("_t") or n.endswith("_a") or n.endswith("_r"))
    cols_to_convert = [col for col in df.columns if looks_like_percent_col(str(col))]
    def to_percent_str(x):
        if pd.isna(x): return x
        s = str(x).strip().replace(" ", "").replace(",", ".")
        if s.endswith("%"):
            try:
                val = float(s[:-1])
                return f"{round(val, 2)}%"
            except:
                return s
        try:
            v = float(s)
            return f"{round(v*100 if v <= 1.5 else v, 2)}%"
        except:
            return x
    for col in cols_to_convert:
        df[col] = df[col].apply(to_percent_str)
    return df

def add_numeric_percent_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    def looks_like_percent_col(name: str) -> bool:
        n = name.strip().lower()
        return ("target" in n or "actual" in n or "rating" in n or "final score" in n
                or n.endswith("_t") or n.endswith("_a") or n.endswith("_r"))
    for col in df.columns:
        if not looks_like_percent_col(str(col)):
            continue
        s = df[col].astype(str).str.strip().str.replace(" ", "", regex=False).str.replace(",", ".", regex=False)
        s = s.str.replace("%", "", regex=False)
        num = pd.to_numeric(s, errors="coerce")
        if num.notna().any():
            num = num.where(num > 1.5, num * 100)
            df[f"{col}_num"] = num
    return df



# ---- Cached loaders (Associates) ----
@st.cache_data(ttl=3600, show_spinner=False)
def load_history_cached() -> pd.DataFrame:
    try:
        return pd.read_excel(HISTORY_FILE)
    except Exception:
        # Return empty frame with expected columns if first run
        return pd.DataFrame(columns=[
            "id","filename","saved_path","uploader","upload_dt","reporting_month",
            "rows_count","source_url","status","message","active","superseded_by","validation_status"
        ])

@st.cache_data(ttl=3600, show_spinner=False)
def load_combined_cached() -> pd.DataFrame:
    try:
        return pd.read_excel(COMBINED_FILE)
    except Exception:
        try:
            return pd.read_csv(COMBINED_FILE_CSV)
        except Exception:
            return pd.DataFrame(columns=["Attachment ID"])


# ---- Cached loaders (BA) ----
@st.cache_data(ttl=3600, show_spinner=False)
def ba_load_history_cached() -> pd.DataFrame:
    try:
        return pd.read_excel(BA_HISTORY_FILE)
    except Exception:
        return pd.DataFrame(columns=[
            "id","filename","saved_path","uploader","upload_dt","reporting_month",
            "rows_count","source_url","status","message","active","superseded_by","validation_status"
        ])

@st.cache_data(ttl=3600, show_spinner=False)
def ba_load_combined_cached() -> pd.DataFrame:
    try:
        return pd.read_excel(BA_COMBINED_FILE)
    except Exception:
        try:
            return pd.read_csv(BA_COMBINED_FILE_CSV)
        except Exception:
            return pd.DataFrame(columns=["Attachment ID"])


# ---- Cached loaders (PE) ----
@st.cache_data(ttl=3600, show_spinner=False)
def pe_load_history_cached() -> pd.DataFrame:
    try:
        return pd.read_excel(PE_HISTORY_FILE)
    except Exception:
        return pd.DataFrame(columns=[
            "id","filename","saved_path","uploader","upload_dt","reporting_month",
            "rows_count","source_url","status","message","active","superseded_by","validation_status"
        ])

@st.cache_data(ttl=3600, show_spinner=False)
def pe_load_combined_cached() -> pd.DataFrame:
    try:
        return pd.read_excel(PE_COMBINED_FILE)
    except Exception:
        try:
            return pd.read_csv(PE_COMBINED_FILE_CSV)
        except Exception:
            return pd.DataFrame(columns=["Attachment ID"])


# ---- Cached loaders (TL) ----
@st.cache_data(ttl=3600, show_spinner=False)
def tl_load_history_cached() -> pd.DataFrame:
    try:
        return pd.read_excel(TL_HISTORY_FILE)
    except Exception:
        return pd.DataFrame(columns=[
            "id","filename","saved_path","uploader","upload_dt","reporting_month",
            "rows_count","source_url","status","message","active","superseded_by","validation_status"
        ])

@st.cache_data(ttl=3600, show_spinner=False)
def tl_load_combined_cached() -> pd.DataFrame:
    try:
        return pd.read_excel(TL_COMBINED_FILE)
    except Exception:
        try:
            return pd.read_csv(TL_COMBINED_FILE_CSV)
        except Exception:
            return pd.DataFrame(columns=["Attachment ID"])


# ---- Cached loaders (PL) ----
@st.cache_data(ttl=3600, show_spinner=False)
def pl_load_history_cached() -> pd.DataFrame:
    try:
        return pd.read_excel(PL_HISTORY_FILE)
    except Exception:
        return pd.DataFrame(columns=[
            "id","filename","saved_path","uploader","upload_dt","reporting_month",
            "rows_count","source_url","status","message","active","superseded_by","validation_status"
        ])

@st.cache_data(ttl=3600, show_spinner=False)
def pl_load_combined_cached() -> pd.DataFrame:
    try:
        return pd.read_excel(PL_COMBINED_FILE)
    except Exception:
        try:
            return pd.read_csv(PL_COMBINED_FILE_CSV)
        except Exception:
            return pd.DataFrame(columns=["Attachment ID"])


# ---- Cached transforms ----
@st.cache_data(ttl=3600, show_spinner=False)
def convert_percentage_cached(df: pd.DataFrame) -> pd.DataFrame:
    return convert_percentage_columns(df)

@st.cache_data(ttl=3600, show_spinner=False)
def add_numeric_cached(df: pd.DataFrame) -> pd.DataFrame:
    return add_numeric_percent_columns(df)


def invalidate_data_caches():
    # Clear all data caches (global)
    st.cache_data.clear()


# -------------------------------------
# Month normalization for filtering (YYYY-MM)
# -------------------------------------
def _to_month_str_series(df: pd.DataFrame) -> pd.Series:
    candidates = [c for c in ["Month", "Reporting Month", "Report Month", "Date"] if c in df.columns]
    if not candidates:
        return pd.Series(dtype="object", index=df.index)
    s = df[candidates[0]]
    dt_series = pd.to_datetime(s, errors="coerce")
    out = dt_series.dt.strftime("%Y-%m")
    if out.isna().all():
        out = s.astype(str).str.strip()
    return out

# -------------------------------------
# Display Cleaner (drop 'Unnamed...' & fully empty columns)
# -------------------------------------
def _is_empty_col(s: pd.Series) -> bool:
    return s.isna().all() or s.astype(str).str.strip().isin(["", "None", "nan", "NaT"]).all()

def clean_dataframe_for_display(df: pd.DataFrame, hide_cols: bool) -> pd.DataFrame:
    if df is None or df.empty or not hide_cols:
        return df
    unnamed_mask = pd.Series(df.columns).str.contains(r"^Unnamed", case=False, na=False)
    df = df.loc[:, ~unnamed_mask.values]
    empty_cols = [c for c in df.columns if _is_empty_col(df[c])]
    return df.drop(columns=empty_cols) if empty_cols else df

# -------------------------------------
# Excel size limits handling
# -------------------------------------
def exceeds_excel_limits(df):
    try:
        r, c = df.shape
        return (r > EXCEL_MAX_ROWS) or (c > EXCEL_MAX_COLS)
    except Exception:
        return False

# -------------------------------------
# Audit Logging
# -------------------------------------
def log_audit(action, attachment_id, filename, user):
    audit_df = pd.read_excel(AUDIT_LOG_FILE)
    audit_df = pd.concat([audit_df, pd.DataFrame([{
        "timestamp": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "action": action, "attachment_id": attachment_id, "filename": filename, "performed_by": user
    }])], ignore_index=True)
    audit_df.to_excel(AUDIT_LOG_FILE, index=False)

# BA audit
def ba_log_audit(action, attachment_id, filename, user):
    audit_df = pd.read_excel(BA_AUDIT_LOG_FILE)
    audit_df = pd.concat([audit_df, pd.DataFrame([{
        "timestamp": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "action": action, "attachment_id": attachment_id, "filename": filename, "performed_by": user
    }])], ignore_index=True)
    audit_df.to_excel(BA_AUDIT_LOG_FILE, index=False)


# PE audit
def pe_log_audit(action, attachment_id, filename, user):
    audit_df = pd.read_excel(PE_AUDIT_LOG_FILE)
    audit_df = pd.concat([audit_df, pd.DataFrame([{
        "timestamp": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "action": action, "attachment_id": attachment_id, "filename": filename, "performed_by": user
    }])], ignore_index=True)
    audit_df.to_excel(PE_AUDIT_LOG_FILE, index=False)


# TL audit
def tl_log_audit(action, attachment_id, filename, user):
    audit_df = pd.read_excel(TL_AUDIT_LOG_FILE)
    audit_df = pd.concat([audit_df, pd.DataFrame([{
        "timestamp": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "action": action, "attachment_id": attachment_id, "filename": filename, "performed_by": user
    }])], ignore_index=True)
    audit_df.to_excel(TL_AUDIT_LOG_FILE, index=False)

# PL audit
def pl_log_audit(action, attachment_id, filename, user):
    audit_df = pd.read_excel(PL_AUDIT_LOG_FILE)
    audit_df = pd.concat([audit_df, pd.DataFrame([{
        "timestamp": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "action": action, "attachment_id": attachment_id, "filename": filename, "performed_by": user
    }])], ignore_index=True)
    audit_df.to_excel(PL_AUDIT_LOG_FILE, index=False)


# -------------------------------------
# Monthly Feedback helpers (upsert)
# -------------------------------------
def load_feedback():
    try:
        return pd.read_excel(FEEDBACK_FILE)
    except Exception:
        return pd.DataFrame(columns=["Domain ID","Name","Month","Team Lead","Feedback","timestamp","entered_by"])

def upsert_feedback(domain_id: str, name: str, month: str, team_lead: str, feedback: str, entered_by: str):
    fb = load_feedback()
    # Remove existing for same Domain ID + Month (edit semantics)
    mask = (fb["Domain ID"].astype(str) == str(domain_id)) & (fb["Month"].astype(str) == str(month))
    fb = fb[~mask]
    new_row = {
        "Domain ID": domain_id, "Name": name, "Month": month, "Team Lead": team_lead,
        "Feedback": feedback, "timestamp": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "entered_by": entered_by
    }
    fb = pd.concat([fb, pd.DataFrame([new_row])], ignore_index=True)
    fb.to_excel(FEEDBACK_FILE, index=False)
    return True


# BA feedback
def ba_load_feedback():
    try:
        return pd.read_excel(BA_FEEDBACK_FILE)
    except Exception:
        return pd.DataFrame(columns=["Domain ID","Name","Month","Team Lead","Feedback","timestamp","entered_by"])

def ba_upsert_feedback(domain_id: str, name: str, month: str, team_lead: str, feedback: str, entered_by: str):
    fb = ba_load_feedback()
    mask = (fb["Domain ID"].astype(str) == str(domain_id)) & (fb["Month"].astype(str) == str(month))
    fb = fb[~mask]
    new_row = {
        "Domain ID": domain_id, "Name": name, "Month": month, "Team Lead": team_lead,
        "Feedback": feedback, "timestamp": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "entered_by": entered_by
    }
    fb = pd.concat([fb, pd.DataFrame([new_row])], ignore_index=True)
    fb.to_excel(BA_FEEDBACK_FILE, index=False)
    return True

# PE feedback
def pe_load_feedback():
    try:
        return pd.read_excel(PE_FEEDBACK_FILE)
    except Exception:
        return pd.DataFrame(columns=["Domain ID","Name","Month","Team Lead","Feedback","timestamp","entered_by"])

def pe_upsert_feedback(domain_id: str, name: str, month: str, team_lead: str, feedback: str, entered_by: str):
    fb = pe_load_feedback()
    mask = (fb["Domain ID"].astype(str) == str(domain_id)) & (fb["Month"].astype(str) == str(month))
    fb = fb[~mask]
    new_row = {
        "Domain ID": domain_id, "Name": name, "Month": month, "Team Lead": team_lead,
        "Feedback": feedback, "timestamp": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "entered_by": entered_by
    }
    fb = pd.concat([fb, pd.DataFrame([new_row])], ignore_index=True)
    fb.to_excel(PE_FEEDBACK_FILE, index=False)
    return True

# TL feedback
def tl_load_feedback():
    try:
        return pd.read_excel(TL_FEEDBACK_FILE)
    except Exception:
        return pd.DataFrame(columns=["Domain ID","Name","Month","Team Lead","Feedback","timestamp","entered_by"])

def tl_upsert_feedback(domain_id: str, name: str, month: str, team_lead: str, feedback: str, entered_by: str):
    fb = tl_load_feedback()
    mask = (fb["Domain ID"].astype(str) == str(domain_id)) & (fb["Month"].astype(str) == str(month))
    fb = fb[~mask]
    new_row = {
        "Domain ID": domain_id, "Name": name, "Month": month, "Team Lead": team_lead,
        "Feedback": feedback, "timestamp": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "entered_by": entered_by
    }
    fb = pd.concat([fb, pd.DataFrame([new_row])], ignore_index=True)
    fb.to_excel(TL_FEEDBACK_FILE, index=False)
    return True

# PL feedback
def pl_load_feedback():
    try:
        return pd.read_excel(PL_FEEDBACK_FILE)
    except Exception:
        return pd.DataFrame(columns=["Domain ID","Name","Month","Team Lead","Feedback","timestamp","entered_by"])

def pl_upsert_feedback(domain_id: str, name: str, month: str, team_lead: str, feedback: str, entered_by: str):
    fb = pl_load_feedback()
    mask = (fb["Domain ID"].astype(str) == str(domain_id)) & (fb["Month"].astype(str) == str(month))
    fb = fb[~mask]
    new_row = {
        "Domain ID": domain_id, "Name": name, "Month": month, "Team Lead": team_lead,
        "Feedback": feedback, "timestamp": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "entered_by": entered_by
    }
    fb = pd.concat([fb, pd.DataFrame([new_row])], ignore_index=True)
    fb.to_excel(PL_FEEDBACK_FILE, index=False)
    return True


# -------------------------------------
# Monthly Metrics table (mirrors filtered table for first set of columns)
# -------------------------------------
def monthly_metrics_table(latest_df: pd.DataFrame, report_month: str, group_by: str = "Domain ID") -> pd.DataFrame:
    """
    Build Associate Monthly Metrics directly from the filtered view.
    - Columns up to Rank mirror the filtered table exactly (no recompute).
    - Merge Team Lead comments by Domain ID + normalized Month (YYYY-MM).
    - Display 'Month' remains exactly as in the filtered table.
    """
    desired_cols = [
        "Domain ID","Function","Function Lead","Team Lead","Designation","Name",
        "Month","Final Score","Rank",
        "Monthly feedback/feedforward","Feedback timestamp"
    ]
    if latest_df is None or latest_df.empty:
        return pd.DataFrame(columns=desired_cols)
    df = latest_df.copy()
    month_norm = _to_month_str_series(df)
    if month_norm.empty or month_norm.isna().all():
        month_norm = pd.Series([str(report_month)] * len(df), index=df.index)
    df["_month_norm_merge"] = month_norm.astype(str)
    fb = load_feedback()
    fb_latest = (
        fb.sort_values("timestamp", ascending=False)
          .drop_duplicates(subset=["Domain ID","Month"], keep="first")
          .rename(columns={"Feedback": "Monthly feedback/feedforward"})
    )
    df = df.merge(
        fb_latest[["Domain ID","Month","Monthly feedback/feedforward","timestamp"]],
        left_on=["Domain ID","_month_norm_merge"],
        right_on=["Domain ID","Month"],
        how="left"
    )
    df = df.rename(columns={"timestamp": "Feedback timestamp"})
    df = df.drop(columns=["Month_y","_month_norm_merge"], errors="ignore")
    df = df.rename(columns={"Month_x": "Month"})
    present = [c for c in desired_cols if c in df.columns]
    out = df[present].copy()
    if "Rank" not in out.columns:
        if "Rank" in df.columns:
            out["Rank"] = df["Rank"]
        elif "Final Score_num" in df.columns:
            out["Rank"] = df["Final Score_num"].rank(method="dense", ascending=False).astype("Int64")
        else:
            if "Final Score" in df.columns:
                s = df["Final Score"].astype(str).str.replace('%','', regex=False).str.replace(',', '.', regex=False)
                num = pd.to_numeric(s, errors='coerce')
                out["Rank"] = num.rank(method="dense", ascending=False).astype("Int64")
    return out

# BA Monthly metrics
def monthly_metrics_table_ba(latest_df: pd.DataFrame, report_month: str, group_by: str = "Domain ID") -> pd.DataFrame:
    desired_cols = [
        "Domain ID","Function","Function Lead","Team Lead","Designation","Name",
        "Month","Final Score","Rank",
        "Monthly feedback/feedforward","Feedback timestamp"
    ]
    if latest_df is None or latest_df.empty:
        return pd.DataFrame(columns=desired_cols)
    df = latest_df.copy()
    month_norm = _to_month_str_series(df)
    if month_norm.empty or month_norm.isna().all():
        month_norm = pd.Series([str(report_month)] * len(df), index=df.index)
    df["_month_norm_merge"] = month_norm.astype(str)
    fb = ba_load_feedback()
    fb_latest = (
        fb.sort_values("timestamp", ascending=False)
          .drop_duplicates(subset=["Domain ID","Month"], keep="first")
          .rename(columns={"Feedback": "Monthly feedback/feedforward"})
    )
    df = df.merge(
        fb_latest[["Domain ID","Month","Monthly feedback/feedforward","timestamp"]],
        left_on=["Domain ID","_month_norm_merge"],
        right_on=["Domain ID","Month"],
        how="left"
    )
    df = df.rename(columns={"timestamp": "Feedback timestamp"})
    df = df.drop(columns=["Month_y","_month_norm_merge"], errors="ignore")
    df = df.rename(columns={"Month_x": "Month"})
    present = [c for c in desired_cols if c in df.columns]
    out = df[present].copy()
    if "Rank" not in out.columns:
        if "Rank" in df.columns:
            out["Rank"] = df["Rank"]
        elif "Final Score_num" in df.columns:
            out["Rank"] = df["Final Score_num"].rank(method="dense", ascending=False).astype("Int64")
        else:
            if "Final Score" in df.columns:
                s = df["Final Score"].astype(str).str.replace('%','', regex=False).str.replace(',', '.', regex=False)
                num = pd.to_numeric(s, errors='coerce')
                out["Rank"] = num.rank(method="dense", ascending=False).astype("Int64")
    return out

# PE Monthly metrics
def monthly_metrics_table_pe(latest_df: pd.DataFrame, report_month: str, group_by: str = "Domain ID") -> pd.DataFrame:
    desired_cols = [
        "Domain ID","Function","Function Lead","Team Lead","Designation","Name",
        "Month","Final Score","Rank",
        "Monthly feedback/feedforward","Feedback timestamp"
    ]
    if latest_df is None or latest_df.empty:
        return pd.DataFrame(columns=desired_cols)
    df = latest_df.copy()
    month_norm = _to_month_str_series(df)
    if month_norm.empty or month_norm.isna().all():
        month_norm = pd.Series([str(report_month)] * len(df), index=df.index)
    df["_month_norm_merge"] = month_norm.astype(str)
    fb = pe_load_feedback()
    fb_latest = (
        fb.sort_values("timestamp", ascending=False)
          .drop_duplicates(subset=["Domain ID","Month"], keep="first")
          .rename(columns={"Feedback": "Monthly feedback/feedforward"})
    )
    df = df.merge(
        fb_latest[["Domain ID","Month","Monthly feedback/feedforward","timestamp"]],
        left_on=["Domain ID","_month_norm_merge"],
        right_on=["Domain ID","Month"],
        how="left"
    )
    df = df.rename(columns={"timestamp": "Feedback timestamp"})
    df = df.drop(columns=["Month_y","_month_norm_merge"], errors="ignore")
    df = df.rename(columns={"Month_x": "Month"})
    present = [c for c in desired_cols if c in df.columns]
    out = df[present].copy()
    if "Rank" not in out.columns:
        if "Rank" in df.columns:
            out["Rank"] = df["Rank"]
        elif "Final Score_num" in df.columns:
            out["Rank"] = df["Final Score_num"].rank(method="dense", ascending=False).astype("Int64")
        else:
            if "Final Score" in df.columns:
                s = df["Final Score"].astype(str).str.replace('%','', regex=False).str.replace(',', '.', regex=False)
                num = pd.to_numeric(s, errors='coerce')
                out["Rank"] = num.rank(method="dense", ascending=False).astype("Int64")
    return out

# TL Monthly metrics
def monthly_metrics_table_tl(latest_df: pd.DataFrame, report_month: str, group_by: str = "Domain ID") -> pd.DataFrame:
    desired_cols = [
        "Domain ID","Function","Function Lead","Team Lead","Designation","Name",
        "Month","Final Score","Rank",
        "Monthly feedback/feedforward","Feedback timestamp"
    ]
    if latest_df is None or latest_df.empty:
        return pd.DataFrame(columns=desired_cols)
    df = latest_df.copy()
    month_norm = _to_month_str_series(df)
    if month_norm.empty or month_norm.isna().all():
        month_norm = pd.Series([str(report_month)] * len(df), index=df.index)
    df["_month_norm_merge"] = month_norm.astype(str)
    fb = tl_load_feedback()
    fb_latest = (
        fb.sort_values("timestamp", ascending=False)
          .drop_duplicates(subset=["Domain ID","Month"], keep="first")
          .rename(columns={"Feedback": "Monthly feedback/feedforward"})
    )
    df = df.merge(
        fb_latest[["Domain ID","Month","Monthly feedback/feedforward","timestamp"]],
        left_on=["Domain ID","_month_norm_merge"],
        right_on=["Domain ID","Month"],
        how="left"
    )
    df = df.rename(columns={"timestamp": "Feedback timestamp"})
    df = df.drop(columns=["Month_y","_month_norm_merge"], errors="ignore")
    df = df.rename(columns={"Month_x": "Month"})
    present = [c for c in desired_cols if c in df.columns]
    out = df[present].copy()
    if "Rank" not in out.columns:
        if "Rank" in df.columns:
            out["Rank"] = df["Rank"]
        elif "Final Score_num" in df.columns:
            out["Rank"] = df["Final Score_num"].rank(method="dense", ascending=False).astype("Int64")
        else:
            if "Final Score" in df.columns:
                s = df["Final Score"].astype(str).str.replace('%','', regex=False).str.replace(',', '.', regex=False)
                num = pd.to_numeric(s, errors='coerce')
                out["Rank"] = num.rank(method="dense", ascending=False).astype("Int64")
    return out


# PL Monthly metrics
def monthly_metrics_table_pl(latest_df: pd.DataFrame, report_month: str, group_by: str = "Domain ID") -> pd.DataFrame:
    desired_cols = [
        "Domain ID","Function","Function Lead","Team Lead","Designation","Name",
        "Month","Final Score","Rank",
        "Monthly feedback/feedforward","Feedback timestamp"
    ]
    if latest_df is None or latest_df.empty:
        return pd.DataFrame(columns=desired_cols)
    df = latest_df.copy()
    month_norm = _to_month_str_series(df)
    if month_norm.empty or month_norm.isna().all():
        month_norm = pd.Series([str(report_month)] * len(df), index=df.index)
    df["_month_norm_merge"] = month_norm.astype(str)
    fb = pl_load_feedback()
    fb_latest = (
        fb.sort_values("timestamp", ascending=False)
          .drop_duplicates(subset=["Domain ID","Month"], keep="first")
          .rename(columns={"Feedback": "Monthly feedback/feedforward"})
    )
    df = df.merge(
        fb_latest[["Domain ID","Month","Monthly feedback/feedforward","timestamp"]],
        left_on=["Domain ID","_month_norm_merge"],
        right_on=["Domain ID","Month"],
        how="left"
    )
    df = df.rename(columns={"timestamp": "Feedback timestamp"})
    df = df.drop(columns=["Month_y","_month_norm_merge"], errors="ignore")
    df = df.rename(columns={"Month_x": "Month"})
    present = [c for c in desired_cols if c in df.columns]
    out = df[present].copy()
    if "Rank" not in out.columns:
        if "Rank" in df.columns:
            out["Rank"] = df["Rank"]
        elif "Final Score_num" in df.columns:
            out["Rank"] = df["Final Score_num"].rank(method="dense", ascending=False).astype("Int64")
        else:
            if "Final Score" in df.columns:
                s = df["Final Score"].astype(str).str.replace('%','', regex=False).str.replace(',', '.', regex=False)
                num = pd.to_numeric(s, errors='coerce')
                out["Rank"] = num.rank(method="dense", ascending=False).astype("Int64")
    return out



# -------------------------------------
# Invalidation & Cleanup / Restore
# -------------------------------------
def mark_invalid_and_cleanup(attachment_id, user):
    history_df = load_history()
    row = history_df[history_df["id"] == attachment_id]
    if row.empty: return False, "Attachment not found"
    filename = row.iloc[0]["filename"]
    saved_path = row.iloc[0]["saved_path"]
    history_df.loc[history_df["id"] == attachment_id, ["validation_status","active"]] = ["Invalid", False]
    save_history(history_df)
    combined_df = load_combined()
    combined_df = combined_df[combined_df["Attachment ID"] != attachment_id]
    save_combined(combined_df)
    invalidate_data_caches()  # ensure next UI run fetches fresh files
    try: os.remove(saved_path)
    except FileNotFoundError: pass
    log_audit("Invalidation & Cleanup", attachment_id, filename, user)
    return True, f"Attachment {attachment_id} marked invalid, deactivated, and data removed."

def mark_valid_and_rebuild(attachment_id, make_active: bool, user: str):
    history_df = load_history()
    row = history_df[history_df["id"] == attachment_id]
    if row.empty: return False, "Attachment not found"
    filename = row.iloc[0]["filename"]
    saved_path = row.iloc[0]["saved_path"]
    month = row.iloc[0]["reporting_month"]
    if not os.path.exists(saved_path):
        return False, "Saved file not found on disk. Re-upload the Excel to restore."
    try:
        with open(saved_path, "rb") as f:
            file_bytes = f.read()
        data_df = read_excel_bytes(file_bytes)
        data_df = convert_percentage_columns(data_df)
        data_df["Attachment ID"] = attachment_id
        combined_df = load_combined()
        combined_df = combined_df[combined_df["Attachment ID"] != attachment_id]
        combined_df = pd.concat([combined_df, data_df], ignore_index=True)
        save_combined(combined_df)
    except Exception as e:
        return False, f"Failed to rebuild data from saved file: {e}"
    history_df.loc[history_df["id"] == attachment_id, "validation_status"] = "Valid"
    if make_active:
        mask = (history_df["reporting_month"] == month) & (history_df["active"] == True) & (history_df["id"] != attachment_id)
        history_df.loc[mask, "active"] = False
        history_df.loc[history_df["id"] == attachment_id, "active"] = True
        history_df.loc[history_df["id"] == attachment_id, "superseded_by"] = ""
    save_history(history_df)
    action = "Restore Valid (active)" if make_active else "Restore Valid"
    log_audit(action, attachment_id, filename, user)
    msg = f"Attachment {attachment_id} marked Valid and indexes rebuilt."
    if make_active: msg += " It is now the active file for the month."
    return True, msg

# BA invalidation & restore
def ba_load_history():
    return pd.read_excel(BA_HISTORY_FILE)

def ba_save_history(df):
    df.to_excel(BA_HISTORY_FILE, index=False)


def ba_load_combined():
    try:
        return pd.read_excel(BA_COMBINED_FILE)
    except Exception:
        try:
            return pd.read_csv(BA_COMBINED_FILE_CSV)
        except Exception:
            return pd.DataFrame(columns=["Attachment ID"])


def ba_save_combined(df):
    if not exceeds_excel_limits(df):
        df.to_excel(BA_COMBINED_FILE, index=False)
    else:
        df.to_csv(BA_COMBINED_FILE_CSV, index=False)


def ba_mark_invalid_and_cleanup(attachment_id, user):
    history_df = ba_load_history_cached()
    row = history_df[history_df["id"] == attachment_id]
    if row.empty: return False, "Attachment not found"
    filename = row.iloc[0]["filename"]
    saved_path = row.iloc[0]["saved_path"]
    history_df.loc[history_df["id"] == attachment_id, ["validation_status","active"]] = ["Invalid", False]
    ba_save_history(history_df)
    combined_df = ba_load_combined_cached()
    combined_df = combined_df[combined_df["Attachment ID"] != attachment_id]
    ba_save_combined(combined_df)
    invalidate_data_caches()  # ensure next UI run fetches fresh files
    try: os.remove(saved_path)
    except FileNotFoundError: pass
    ba_log_audit("Invalidation & Cleanup", attachment_id, filename, user)
    return True, f"Attachment {attachment_id} marked invalid, deactivated, and data removed."


def ba_mark_valid_and_rebuild(attachment_id, make_active: bool, user: str):
    history_df = ba_load_history_cached()
    row = history_df[history_df["id"] == attachment_id]
    if row.empty: return False, "Attachment not found"
    filename = row.iloc[0]["filename"]
    saved_path = row.iloc[0]["saved_path"]
    month = row.iloc[0]["reporting_month"]
    if not os.path.exists(saved_path):
        return False, "Saved file not found on disk. Re-upload the Excel to restore."
    try:
        with open(saved_path, "rb") as f:
            file_bytes = f.read()
        data_df = read_excel_bytes(file_bytes)
        data_df = convert_percentage_columns(data_df)
        data_df["Attachment ID"] = attachment_id
        combined_df = ba_load_combined_cached()
        combined_df = combined_df[combined_df["Attachment ID"] != attachment_id]
        combined_df = pd.concat([combined_df, data_df], ignore_index=True)
        ba_save_combined(combined_df)
    except Exception as e:
        return False, f"Failed to rebuild data from saved file: {e}"
    history_df.loc[history_df["id"] == attachment_id, "validation_status"] = "Valid"
    if make_active:
        mask = (history_df["reporting_month"] == month) & (history_df["active"] == True) & (history_df["id"] != attachment_id)
        history_df.loc[mask, "active"] = False
        history_df.loc[history_df["id"] == attachment_id, "active"] = True
        history_df.loc[history_df["id"] == attachment_id, "superseded_by"] = ""
    ba_save_history(history_df)
    action = "Restore Valid (active)" if make_active else "Restore Valid"
    ba_log_audit(action, attachment_id, filename, user)
    msg = f"Attachment {attachment_id} marked Valid and indexes rebuilt."
    if make_active: msg += " It is now the active file for the month."
    return True, msg



# PE invalidation & restore
def pe_load_history():
    return pd.read_excel(PE_HISTORY_FILE)

def pe_save_history(df):
    df.to_excel(PE_HISTORY_FILE, index=False)


def pe_load_combined():
    try:
        return pd.read_excel(PE_COMBINED_FILE)
    except Exception:
        try:
            return pd.read_csv(PE_COMBINED_FILE_CSV)
        except Exception:
            return pd.DataFrame(columns=["Attachment ID"])


def pe_save_combined(df):
    if not exceeds_excel_limits(df):
        df.to_excel(PE_COMBINED_FILE, index=False)
    else:
        df.to_csv(PE_COMBINED_FILE_CSV, index=False)


def pe_mark_invalid_and_cleanup(attachment_id, user):
    history_df = pe_load_history_cached()
    row = history_df[history_df["id"] == attachment_id]
    if row.empty: return False, "Attachment not found"
    filename = row.iloc[0]["filename"]
    saved_path = row.iloc[0]["saved_path"]
    history_df.loc[history_df["id"] == attachment_id, ["validation_status","active"]] = ["Invalid", False]
    pe_save_history(history_df)
    combined_df = pe_load_combined_cached()
    combined_df = combined_df[combined_df["Attachment ID"] != attachment_id]
    pe_save_combined(combined_df)
    invalidate_data_caches()  # ensure next UI run fetches fresh files
    try: os.remove(saved_path)
    except FileNotFoundError: pass
    pe_log_audit("Invalidation & Cleanup", attachment_id, filename, user)
    return True, f"Attachment {attachment_id} marked invalid, deactivated, and data removed."


def pe_mark_valid_and_rebuild(attachment_id, make_active: bool, user: str):
    history_df = pe_load_history_cached()
    row = history_df[history_df["id"] == attachment_id]
    if row.empty: return False, "Attachment not found"
    filename = row.iloc[0]["filename"]
    saved_path = row.iloc[0]["saved_path"]
    month = row.iloc[0]["reporting_month"]
    if not os.path.exists(saved_path):
        return False, "Saved file not found on disk. Re-upload the Excel to restore."
    try:
        with open(saved_path, "rb") as f:
            file_bytes = f.read()
        data_df = read_excel_bytes(file_bytes)
        data_df = convert_percentage_columns(data_df)
        data_df["Attachment ID"] = attachment_id
        combined_df = pe_load_combined_cached()
        combined_df = combined_df[combined_df["Attachment ID"] != attachment_id]
        combined_df = pd.concat([combined_df, data_df], ignore_index=True)
        pe_save_combined(combined_df)
    except Exception as e:
        return False, f"Failed to rebuild data from saved file: {e}"
    history_df.loc[history_df["id"] == attachment_id, "validation_status"] = "Valid"
    if make_active:
        mask = (history_df["reporting_month"] == month) & (history_df["active"] == True) & (history_df["id"] != attachment_id)
        history_df.loc[mask, "active"] = False
        history_df.loc[history_df["id"] == attachment_id, "active"] = True
        history_df.loc[history_df["id"] == attachment_id, "superseded_by"] = ""
    pe_save_history(history_df)
    action = "Restore Valid (active)" if make_active else "Restore Valid"
    pe_log_audit(action, attachment_id, filename, user)
    msg = f"Attachment {attachment_id} marked Valid and indexes rebuilt."
    if make_active: msg += " It is now the active file for the month."
    return True, msg


# TL invalidation & restore
def tl_load_history():
    return pd.read_excel(TL_HISTORY_FILE)

def tl_save_history(df):
    df.to_excel(TL_HISTORY_FILE, index=False)

def tl_load_combined():
    try:
        return pd.read_excel(TL_COMBINED_FILE)
    except Exception:
        try:
            return pd.read_csv(TL_COMBINED_FILE_CSV)
        except Exception:
            return pd.DataFrame(columns=["Attachment ID"])

def tl_save_combined(df):
    if not exceeds_excel_limits(df):
        df.to_excel(TL_COMBINED_FILE, index=False)
    else:
        df.to_csv(TL_COMBINED_FILE_CSV, index=False)

def tl_mark_invalid_and_cleanup(attachment_id, user):
    history_df = tl_load_history_cached()
    row = history_df[history_df["id"] == attachment_id]
    if row.empty: return False, "Attachment not found"
    filename = row.iloc[0]["filename"]
    saved_path = row.iloc[0]["saved_path"]
    history_df.loc[history_df["id"] == attachment_id, ["validation_status","active"]] = ["Invalid", False]
    tl_save_history(history_df)
    combined_df = tl_load_combined_cached()
    combined_df = combined_df[combined_df["Attachment ID"] != attachment_id]
    tl_save_combined(combined_df)
    invalidate_data_caches()  # ensure next UI run fetches fresh files
    try: os.remove(saved_path)
    except FileNotFoundError: pass
    tl_log_audit("Invalidation & Cleanup", attachment_id, filename, user)
    return True, f"Attachment {attachment_id} marked invalid, deactivated, and data removed."

def tl_mark_valid_and_rebuild(attachment_id, make_active: bool, user: str):
    history_df = tl_load_history_cached()
    row = history_df[history_df["id"] == attachment_id]
    if row.empty: return False, "Attachment not found"
    filename = row.iloc[0]["filename"]
    saved_path = row.iloc[0]["saved_path"]
    month = row.iloc[0]["reporting_month"]
    if not os.path.exists(saved_path):
        return False, "Saved file not found on disk. Re-upload the Excel to restore."
    try:
        with open(saved_path, "rb") as f:
            file_bytes = f.read()
        data_df = read_excel_bytes(file_bytes)
        data_df = convert_percentage_columns(data_df)
        data_df["Attachment ID"] = attachment_id
        combined_df = tl_load_combined_cached()
        combined_df = combined_df[combined_df["Attachment ID"] != attachment_id]
        combined_df = pd.concat([combined_df, data_df], ignore_index=True)
        tl_save_combined(combined_df)
    except Exception as e:
        return False, f"Failed to rebuild data from saved file: {e}"
    history_df.loc[history_df["id"] == attachment_id, "validation_status"] = "Valid"
    if make_active:
        mask = (history_df["reporting_month"] == month) & (history_df["active"] == True) & (history_df["id"] != attachment_id)
        history_df.loc[mask, "active"] = False
        history_df.loc[history_df["id"] == attachment_id, "active"] = True
        history_df.loc[history_df["id"] == attachment_id, "superseded_by"] = ""
    tl_save_history(history_df)
    action = "Restore Valid (active)" if make_active else "Restore Valid"
    tl_log_audit(action, attachment_id, filename, user)
    msg = f"Attachment {attachment_id} marked Valid and indexes rebuilt."
    if make_active: msg += " It is now the active file for the month."
    return True, msg


# PL invalidation & restore
def pl_load_history():
    return pd.read_excel(PL_HISTORY_FILE)

def pl_save_history(df):
    df.to_excel(PL_HISTORY_FILE, index=False)

def pl_load_combined():
    try:
        return pd.read_excel(PL_COMBINED_FILE)
    except Exception:
        try:
            return pd.read_csv(PL_COMBINED_FILE_CSV)
        except Exception:
            return pd.DataFrame(columns=["Attachment ID"])

def pl_save_combined(df):
    if not exceeds_excel_limits(df):
        df.to_excel(PL_COMBINED_FILE, index=False)
    else:
        df.to_csv(PL_COMBINED_FILE_CSV, index=False)

def pl_mark_invalid_and_cleanup(attachment_id, user):
    history_df = pl_load_history_cached()
    row = history_df[history_df["id"] == attachment_id]
    if row.empty: return False, "Attachment not found"
    filename = row.iloc[0]["filename"]
    saved_path = row.iloc[0]["saved_path"]
    history_df.loc[history_df["id"] == attachment_id, ["validation_status","active"]] = ["Invalid", False]
    pl_save_history(history_df)
    combined_df = pl_load_combined_cached()
    combined_df = combined_df[combined_df["Attachment ID"] != attachment_id]
    pl_save_combined(combined_df)
    invalidate_data_caches()  # ensure next UI run fetches fresh files
    try: os.remove(saved_path)
    except FileNotFoundError: pass
    pl_log_audit("Invalidation & Cleanup", attachment_id, filename, user)
    return True, f"Attachment {attachment_id} marked invalid, deactivated, and data removed."

def pl_mark_valid_and_rebuild(attachment_id, make_active: bool, user: str):
    history_df = pl_load_history_cached()
    row = history_df[history_df["id"] == attachment_id]
    if row.empty: return False, "Attachment not found"
    filename = row.iloc[0]["filename"]
    saved_path = row.iloc[0]["saved_path"]
    month = row.iloc[0]["reporting_month"]
    if not os.path.exists(saved_path):
        return False, "Saved file not found on disk. Re-upload the Excel to restore."
    try:
        with open(saved_path, "rb") as f:
            file_bytes = f.read()
        data_df = read_excel_bytes(file_bytes)
        data_df = convert_percentage_columns(data_df)
        data_df["Attachment ID"] = attachment_id
        combined_df = pl_load_combined_cached()
        combined_df = combined_df[combined_df["Attachment ID"] != attachment_id]
        combined_df = pd.concat([combined_df, data_df], ignore_index=True)
        pl_save_combined(combined_df)
    except Exception as e:
        return False, f"Failed to rebuild data from saved file: {e}"
    history_df.loc[history_df["id"] == attachment_id, "validation_status"] = "Valid"
    if make_active:
        mask = (history_df["reporting_month"] == month) & (history_df["active"] == True) & (history_df["id"] != attachment_id)
        history_df.loc[mask, "active"] = False
        history_df.loc[history_df["id"] == attachment_id, "active"] = True
        history_df.loc[history_df["id"] == attachment_id, "superseded_by"] = ""
    pl_save_history(history_df)
    action = "Restore Valid (active)" if make_active else "Restore Valid"
    pl_log_audit(action, attachment_id, filename, user)
    msg = f"Attachment {attachment_id} marked Valid and indexes rebuilt."
    if make_active: msg += " It is now the active file for the month."
    return True, msg




# -------------------------------------
# Upload Processing
# -------------------------------------
def process_upload(name, file_bytes, uploader, source_url=None):
    if len(file_bytes) > MAX_UPLOAD_MB*1024*1024:
        return False, f"File exceeds {MAX_UPLOAD_MB}MB", None
    try:
        data_df = read_excel_bytes(file_bytes)
    except ValueError as e:
        return False, str(e), None
    missing_data = validate_required_columns(data_df)
    if missing_data:
        return False, f"Missing required columns in Data sheet: {', '.join(missing_data)}", None
    data_df = convert_percentage_columns(data_df)
    month = safe_month_from_columns(data_df)
    attach_id = str(uuid.uuid4())
    path = derive_saved_path(month, name)
    with open(path, "wb") as f:
        f.write(file_bytes)
    history_df = load_history()
    history_df = supersede_existing_month(month, attach_id, history_df)
    history_df = pd.concat([history_df, pd.DataFrame([{
        "id": attach_id, "filename": name, "saved_path": path, "uploader": uploader,
        "upload_dt": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "reporting_month": month,
        "rows_count": len(data_df), "source_url": source_url or "", "status": "success",
        "message": "", "active": True, "superseded_by": "", "validation_status": "Valid"
    }])], ignore_index=True)
    save_history(history_df)
    data_df["Attachment ID"] = attach_id
    combined_df = load_combined()
    combined_df = pd.concat([combined_df, data_df], ignore_index=True)
    save_combined(combined_df)
    invalidate_data_caches()  # ensure next UI run fetches fresh files
    return True, f"Uploaded and processed for month {month}.", data_df.head(20)

# BA Upload Processing
def ba_process_upload(name, file_bytes, uploader, source_url=None):
    if len(file_bytes) > MAX_UPLOAD_MB*1024*1024:
        return False, f"File exceeds {MAX_UPLOAD_MB}MB", None
    try:
        data_df = read_excel_bytes(file_bytes)
    except ValueError as e:
        return False, str(e), None
    missing_data = validate_required_columns(data_df)
    if missing_data:
        return False, f"Missing required columns in Data sheet: {', '.join(missing_data)}", None
    data_df = convert_percentage_columns(data_df)
    month = safe_month_from_columns(data_df)
    attach_id = str(uuid.uuid4())
    path = os.path.join(BA_ATTACHMENTS_DIR, f"{month}_{name.replace('/', '_').replace(chr(92), '_')}")
    with open(path, "wb") as f:
        f.write(file_bytes)
    history_df = ba_load_history_cached()
    mask = (history_df["reporting_month"] == month) & (history_df["active"] == True)
    for idx in history_df[mask].index:
        try: os.remove(history_df.at[idx, "saved_path"])
        except: pass
        history_df.at[idx, "active"] = False
        history_df.at[idx, "superseded_by"] = attach_id
    history_df = pd.concat([history_df, pd.DataFrame([{
        "id": attach_id, "filename": name, "saved_path": path, "uploader": uploader,
        "upload_dt": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "reporting_month": month,
        "rows_count": len(data_df), "source_url": source_url or "", "status": "success",
        "message": "", "active": True, "superseded_by": "", "validation_status": "Valid"
    }])], ignore_index=True)
    ba_save_history(history_df)
    data_df["Attachment ID"] = attach_id
    combined_df = ba_load_combined_cached()
    combined_df = pd.concat([combined_df, data_df], ignore_index=True)
    ba_save_combined(combined_df)
    invalidate_data_caches()  # ensure next UI run fetches fresh files
    return True, f"Uploaded and processed for month {month}.", data_df.head(20)

# PE Upload Processing
def pe_process_upload(name, file_bytes, uploader, source_url=None):
    if len(file_bytes) > MAX_UPLOAD_MB*1024*1024:
        return False, f"File exceeds {MAX_UPLOAD_MB}MB", None
    try:
        data_df = read_excel_bytes(file_bytes)
    except ValueError as e:
        return False, str(e), None
    missing_data = validate_required_columns(data_df)
    if missing_data:
        return False, f"Missing required columns in Data sheet: {', '.join(missing_data)}", None
    data_df = convert_percentage_columns(data_df)
    month = safe_month_from_columns(data_df)
    attach_id = str(uuid.uuid4())
    path = os.path.join(PE_ATTACHMENTS_DIR, f"{month}_{name.replace('/', '_').replace(chr(92), '_')}")
    with open(path, "wb") as f:
        f.write(file_bytes)
    history_df = pe_load_history_cached()
    # Supersede any active file for the same month
    mask = (history_df["reporting_month"] == month) & (history_df["active"] == True)
    for idx in history_df[mask].index:
        try: os.remove(history_df.at[idx, "saved_path"])
        except: pass
        history_df.at[idx, "active"] = False
        history_df.at[idx, "superseded_by"] = attach_id
    history_df = pd.concat([history_df, pd.DataFrame([{
        "id": attach_id, "filename": name, "saved_path": path, "uploader": uploader,
        "upload_dt": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "reporting_month": month,
        "rows_count": len(data_df), "source_url": source_url or "", "status": "success",
        "message": "", "active": True, "superseded_by": "", "validation_status": "Valid"
    }])], ignore_index=True)
    pe_save_history(history_df)
    data_df["Attachment ID"] = attach_id
    combined_df = pe_load_combined_cached()
    combined_df = pd.concat([combined_df, data_df], ignore_index=True)
    pe_save_combined(combined_df)
    invalidate_data_caches()  # ensure next UI run fetches fresh files
    return True, f"Uploaded and processed for month {month}.", data_df.head(20)

# TL Upload Processing
def tl_process_upload(name, file_bytes, uploader, source_url=None):
    if len(file_bytes) > MAX_UPLOAD_MB*1024*1024:
        return False, f"File exceeds {MAX_UPLOAD_MB}MB", None
    try:
        data_df = read_excel_bytes(file_bytes)
    except ValueError as e:
        return False, str(e), None
    missing_data = validate_required_columns(data_df)
    if missing_data:
        return False, f"Missing required columns in Data sheet: {', '.join(missing_data)}", None
    data_df = convert_percentage_columns(data_df)
    month = safe_month_from_columns(data_df)
    attach_id = str(uuid.uuid4())
    path = os.path.join(TL_ATTACHMENTS_DIR, f"{month}_{name.replace('/', '_').replace(chr(92), '_')}")
    with open(path, "wb") as f:
        f.write(file_bytes)
    history_df = tl_load_history_cached()
    # Supersede any active file for the same month
    mask = (history_df["reporting_month"] == month) & (history_df["active"] == True)
    for idx in history_df[mask].index:
        try: os.remove(history_df.at[idx, "saved_path"])
        except: pass
        history_df.at[idx, "active"] = False
        history_df.at[idx, "superseded_by"] = attach_id
    history_df = pd.concat([history_df, pd.DataFrame([{
        "id": attach_id, "filename": name, "saved_path": path, "uploader": uploader,
        "upload_dt": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "reporting_month": month,
        "rows_count": len(data_df), "source_url": source_url or "", "status": "success",
        "message": "", "active": True, "superseded_by": "", "validation_status": "Valid"
    }])], ignore_index=True)
    tl_save_history(history_df)
    data_df["Attachment ID"] = attach_id
    combined_df = tl_load_combined_cached()
    combined_df = pd.concat([combined_df, data_df], ignore_index=True)
    tl_save_combined(combined_df)
    invalidate_data_caches()  # ensure next UI run fetches fresh files
    return True, f"Uploaded and processed for month {month}.", data_df.head(20)


# PL Upload Processing
def pl_process_upload(name, file_bytes, uploader, source_url=None):
    if len(file_bytes) > MAX_UPLOAD_MB*1024*1024:
        return False, f"File exceeds {MAX_UPLOAD_MB}MB", None
    try:
        data_df = read_excel_bytes(file_bytes)
    except ValueError as e:
        return False, str(e), None
    missing_data = validate_required_columns(data_df)
    if missing_data:
        return False, f"Missing required columns in Data sheet: {', '.join(missing_data)}", None
    data_df = convert_percentage_columns(data_df)
    month = safe_month_from_columns(data_df)
    attach_id = str(uuid.uuid4())
    path = os.path.join(PL_ATTACHMENTS_DIR, f"{month}_{name.replace('/', '_').replace(chr(92), '_')}")
    with open(path, "wb") as f:
        f.write(file_bytes)
    history_df = pl_load_history_cached()
    # Supersede any active file for the same month
    mask = (history_df["reporting_month"] == month) & (history_df["active"] == True)
    for idx in history_df[mask].index:
        try: os.remove(history_df.at[idx, "saved_path"])
        except: pass
        history_df.at[idx, "active"] = False
        history_df.at[idx, "superseded_by"] = attach_id
    history_df = pd.concat([history_df, pd.DataFrame([{
        "id": attach_id, "filename": name, "saved_path": path, "uploader": uploader,
        "upload_dt": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "reporting_month": month,
        "rows_count": len(data_df), "source_url": source_url or "", "status": "success",
        "message": "", "active": True, "superseded_by": "", "validation_status": "Valid"
    }])], ignore_index=True)
    pl_save_history(history_df)
    data_df["Attachment ID"] = attach_id
    combined_df = pl_load_combined_cached()
    combined_df = pd.concat([combined_df, data_df], ignore_index=True)
    pl_save_combined(combined_df)
    invalidate_data_caches()  # ensure next UI run fetches fresh files
    return True, f"Uploaded and processed for month {month}.", data_df.head(20)



# -------------------------------------
# Filtering & Search (shared helpers)
# -------------------------------------
def filter_combined(df, d_ids, funcs, f_leads, t_leads, months=None):
    if d_ids and "Domain ID" in df.columns:
        df = df[df["Domain ID"].astype(str).isin(d_ids)]
    if funcs and "Function" in df.columns:
        df = df[df["Function"].astype(str).isin(funcs)]
    if f_leads and "Function Lead" in df.columns:
        df = df[df["Function Lead"].astype(str).isin(f_leads)]
    if t_leads and "Team Lead" in df.columns:
        df = df[df["Team Lead"].astype(str).isin(t_leads)]
    if months:
        m = _to_month_str_series(df)
        if not m.empty:
            df = df[m.isin(months)]
    return df

def apply_search(df, q):
    q = q.strip()
    if not q:
        return df
    mask = pd.Series(False, index=df.index)
    for col in df.columns:
        try:
            mask = df[col].astype(str).str.contains(q, case=False, na=False)
        except Exception:
            continue
    return df[mask]



# --- Final score band filter (UI only) ---
def apply_final_score_band_filter(df: pd.DataFrame, band: str) -> pd.DataFrame:
    """
    Filters rows by 'Final Score' value bands for UI display:
      - '>= 100'
      - 'Between 90 and 99.99'
      - '< 90'
      - 'All' (no filter)

    Works with both % strings and numeric columns by using Final Score_num.
    """
    if df is None or df.empty or not band or band == "All":
        return df

    tmp = df.copy()

    # Ensure numeric companion column exists
    if "Final Score_num" not in tmp.columns:
        tmp = add_numeric_percent_columns(tmp)

    if "Final Score_num" not in tmp.columns:
        # Gracefully skip if still unavailable
        return df

    if band == ">= 100":
        mask = tmp["Final Score_num"] >= 100
    elif band == "Between 90 and 99.99":
        mask = (tmp["Final Score_num"] >= 90) & (tmp["Final Score_num"] < 100)
    elif band == "< 90":
        mask = tmp["Final Score_num"] < 90
    else:
        # Unknown value -> no-op
        return df

    return tmp[mask]




def make_excel_bytes_from_df(df, hide_cols: bool):
    df = clean_dataframe_for_display(df, hide_cols)
    if exceeds_excel_limits(df):
        buf = io.BytesIO()
        df.to_csv(buf, index=False)
        return buf.getvalue()
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df.to_excel(w, index=False)
    return buf.getvalue()

###Color code the Final score column
def _final_score_to_number(val):
    """
    Robustly converts Final Score values like '97%', '97.0', '97,0', ' 97 % ' to float 97.0.
    If conversion fails, returns None.
    """
    try:
        s = str(val).strip().replace('%', '').replace(' ', '').replace(',', '.')
        return float(s)
    except Exception:
        return None


def style_associates_metrics_df(df: pd.DataFrame):
    """
    Returns a pd.Styler with the Final Score column color-coded for the Associates Monthly view.
    Only colors 'Final Score' if that column exists.
    """
    if df is None or df.empty or ("Final Score" not in df.columns):
        return df

    def _style_series(s: pd.Series):
        styles = []
        for v in s:
            num = _final_score_to_number(v)
            if num is None:
                styles.append('')  # no style for missing/invalid
                continue
            if num >= 100:
                styles.append('background-color: #C6EFCE; color: #1E4620;')  # light green
            elif 90 <= num < 100:
                styles.append('background-color: #FFF3CD; color: #664D03;')  # light yellow
            else:
                styles.append('background-color: #F8D7DA; color: #58151C;')  # light red
        return styles

    # Use Styler.apply on the Final Score column only
    return df.style.apply(_style_series, subset=["Final Score"])


def make_excel_bytes_associates_monthly_metrics(df: pd.DataFrame, hide_cols: bool):
    """
    Creates an Excel bytes payload for the Associates Monthly Metrics table,
    with conditional formatting applied to the 'Final Score' column:
        >= 100 -> green
        90-99.99 -> yellow
        < 90 -> red

    Falls back to CSV if Excel size limits are exceeded.
    """
    # Clean display (respect global toggle)
    df_clean = clean_dataframe_for_display(df, hide_cols)

    # If too wide for Excel, keep parity with existing behavior and return CSV
    if exceeds_excel_limits(df_clean):
        buf = io.BytesIO()
        df_clean.to_csv(buf, index=False)
        return buf.getvalue()

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        sheet_name = "Associates Monthly Metrics"
        df_clean.to_excel(writer, index=False, sheet_name=sheet_name)

        wb = writer.book
        ws = writer.sheets[sheet_name]

        # If 'Final Score' exists, ensure numeric values for CF and apply styles
        if "Final Score" in df_clean.columns:
            col_idx = list(df_clean.columns).index("Final Score") + 1  # 1-based
            col_letter = get_column_letter(col_idx)

            # Convert written cells to numeric (so CellIsRule works), set a friendly number format
            for r in range(2, ws.max_row + 1):  # skip header row (1)
                cell = ws[f"{col_letter}{r}"]
                num = _final_score_to_number(cell.value)
                if num is not None:
                    cell.value = num
                    # Show as e.g., 97.0% (optional). If you prefer just 97.0, use '0.0'
                    cell.number_format = '0.0"%"'

            # Conditional formatting fills
            green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')   # >= 100
            yellow_fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid') # 90-99.99
            red_fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')    # < 90

            data_range = f"{col_letter}2:{col_letter}{ws.max_row}"

            # >= 100 -> green
            ws.conditional_formatting.add(
                data_range,
                CellIsRule(operator='greaterThanOrEqual', formula=['100'], stopIfTrue=False, fill=green_fill)
            )
            # between 90 and 99.99 -> yellow
            ws.conditional_formatting.add(
                data_range,
                CellIsRule(operator='between', formula=['90', '99.99'], stopIfTrue=False, fill=yellow_fill)
            )
            # < 90 -> red
            ws.conditional_formatting.add(
                data_range,
                CellIsRule(operator='lessThan', formula=['90'], stopIfTrue=False, fill=red_fill)
            )

    return buf.getvalue()



def make_excel_bytes_associates_ytd_aggregated(df: pd.DataFrame, hide_cols: bool):
    """
    Excel bytes for YTD Aggregated Associates table with conditional formatting
    on 'Final Score' column:
      >= 100  -> green
      90-99.99 -> yellow
      < 90    -> red
    Falls back to CSV if Excel size limits are exceeded.
    """
    df_clean = clean_dataframe_for_display(df, hide_cols)

    if exceeds_excel_limits(df_clean):
        buf = io.BytesIO()
        df_clean.to_csv(buf, index=False)
        return buf.getvalue()

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        sheet_name = "Associates YTD Aggregated"
        df_clean.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]

        if "Final Score" in df_clean.columns:
            col_idx = list(df_clean.columns).index("Final Score") + 1
            col_letter = get_column_letter(col_idx)

            # Normalize to numeric for CF and apply display format
            for r in range(2, ws.max_row + 1):
                cell = ws[f"{col_letter}{r}"]
                num = _final_score_to_number(cell.value)
                if num is not None:
                    cell.value = num
                # Show as percentage (optional)
                cell.number_format = '0.0"%"'

            green_fill  = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            yellow_fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
            red_fill    = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')

            data_range = f"{col_letter}2:{col_letter}{ws.max_row}"

            # >= 100 -> green
            ws.conditional_formatting.add(
                data_range,
                CellIsRule(operator='greaterThanOrEqual', formula=['100'], stopIfTrue=False, fill=green_fill)
            )
            # between 90 and 99.99 -> yellow
            ws.conditional_formatting.add(
                data_range,
                CellIsRule(operator='between', formula=['90', '99.99'], stopIfTrue=False, fill=yellow_fill)
            )
            # < 90 -> red
            ws.conditional_formatting.add(
                data_range,
                CellIsRule(operator='lessThan', formula=['90'], stopIfTrue=False, fill=red_fill)
            )

    return buf.getvalue()



# Helper to fetch latest active monthly data (Associates)
def get_latest_monthly_data():
    h = load_history()
    active_mask = _coerce_active_bool(h.get("active", pd.Series([], dtype="object")))
    active = h[active_mask]
    if active.empty:
        return None, None, None
    latest_row = active.sort_values("upload_dt", ascending=False).iloc[0]
    latest_id = latest_row["id"]
    combined_df = load_combined()
    latest_data = combined_df[combined_df["Attachment ID"] == latest_id]
    return latest_row, latest_id, latest_data

# Helper to fetch latest active monthly data (BA)
def ba_get_latest_monthly_data():
    h = ba_load_history_cached()
    active_mask = _coerce_active_bool(h.get("active", pd.Series([], dtype="object")))
    active = h[active_mask]
    if active.empty:
        return None, None, None
    latest_row = active.sort_values("upload_dt", ascending=False).iloc[0]
    latest_id = latest_row["id"]
    combined_df = ba_load_combined_cached()
    latest_data = combined_df[combined_df["Attachment ID"] == latest_id]
    return latest_row, latest_id, latest_data

# Helper to fetch latest active monthly data (PE)
def pe_get_latest_monthly_data():
    h = pe_load_history_cached()
    active_mask = _coerce_active_bool(h.get("active", pd.Series([], dtype="object")))
    active = h[active_mask]
    if active.empty:
        return None, None, None
    latest_row = active.sort_values("upload_dt", ascending=False).iloc[0]
    latest_id = latest_row["id"]
    combined_df = pe_load_combined_cached()
    latest_data = combined_df[combined_df["Attachment ID"] == latest_id]
    return latest_row, latest_id, latest_data

# Helper to fetch latest active monthly data (TL)
def tl_get_latest_monthly_data():
    h = tl_load_history_cached()
    active_mask = _coerce_active_bool(h.get("active", pd.Series([], dtype="object")))
    active = h[active_mask]
    if active.empty:
        return None, None, None
    latest_row = active.sort_values("upload_dt", ascending=False).iloc[0]
    latest_id = latest_row["id"]
    combined_df = tl_load_combined_cached()
    latest_data = combined_df[combined_df["Attachment ID"] == latest_id]
    return latest_row, latest_id, latest_data

# Helper to fetch latest active monthly data (PL)
def pl_get_latest_monthly_data():
    h = pl_load_history_cached()
    active_mask = _coerce_active_bool(h.get("active", pd.Series([], dtype="object")))
    active = h[active_mask]
    if active.empty:
        return None, None, None
    latest_row = active.sort_values("upload_dt", ascending=False).iloc[0]
    latest_id = latest_row["id"]
    combined_df = pl_load_combined_cached()
    latest_data = combined_df[combined_df["Attachment ID"] == latest_id]
    return latest_row, latest_id, latest_data



# -------------------------------------
# Visualization helpers
# -------------------------------------
PALETTES = {
    "Blue": ["#1f77b4"],
    "Green": ["#2ca02c"],
    "Orange": ["#ff7f0e"],
    "Purple": ["#9467bd"],
    "Teal": ["#17becf"],
    "Category10": "category10",
    "Tableau10": "tableau10",
}
def enable_altair_theme():
    def _theme():
        return {
            "config": {
                "view": {"continuousHeight": 300, "continuousWidth": 480},
                "axis": {
                    "labelFont": "Arial",
                    "titleFont": "Arial",
                    "labelColor": "#333",
                    "titleColor": "#333",
                    "grid": True,
                    "gridColor": "#eee",
                },
                "legend": {"labelFont": "Arial", "titleFont": "Arial"},
                "bar": {"cornerRadius": 4},
                "mark": {"tooltip": {"content": "encoding"}},
            }
        }
    alt.themes.register("scorecard_theme", _theme)
    alt.themes.enable("scorecard_theme")

def get_numeric_metric_options(df: pd.DataFrame):
    return [c for c in df.columns if c.endswith("_num")]

def aggregate_df(df: pd.DataFrame, dim: str, metric: str, method: str = "mean"):
    if dim not in df.columns or metric not in df.columns:
        return pd.DataFrame(columns=[dim, metric])
    tmp = df.dropna(subset=[dim, metric]).copy()
    if method == "median":
        out = tmp.groupby(dim, as_index=False)[metric].median()
    else:
        out = tmp.groupby(dim, as_index=False)[metric].mean()
    return out

def add_rank_and_topN(agg: pd.DataFrame, dim: str, metric: str, top_n: int = 15, ascending: bool = False):
    if agg.empty:
        return agg
    agg = agg.sort_values(metric, ascending=ascending)
    return agg.head(top_n)

def bar_chart(agg: pd.DataFrame, dim: str, metric: str, title: str, palette: str = "Category10", show_labels: bool = True):
    if agg.empty:
        return alt.Chart(pd.DataFrame())
    color_enc = alt.Color(f"{dim}:N", legend=alt.Legend(title=dim))
    if palette in PALETTES and isinstance(PALETTES[palette], list):
        color_enc = alt.Color(f"{dim}:N", scale=alt.Scale(range=PALETTES[palette]), legend=alt.Legend(title=dim))
    elif palette in PALETTES and isinstance(PALETTES[palette], str):
        color_enc = alt.Color(f"{dim}:N", scale=alt.Scale(scheme=PALETTES[palette]), legend=alt.Legend(title=dim))
    base = alt.Chart(agg).mark_bar().encode(
        x=alt.X(f"{metric}:Q", title=metric.replace("_num", " (%)")),
        y=alt.Y(f"{dim}:N", sort="-x", title=dim),
        color=color_enc,
        tooltip=[alt.Tooltip(f"{dim}:N", title=dim), alt.Tooltip(f"{metric}:Q", format=".1f", title=metric.replace("_num", " (%)"))]
    ).properties(title=title)
    if show_labels:
        text = alt.Chart(agg).mark_text(dx=4, color="#333", align="left").encode(
            y=alt.Y(f"{dim}:N", sort="-x"),
            x=alt.X(f"{metric}:Q"),
            text=alt.Text(f"{metric}:Q", format=".1f")
        )
        return base + text
    return base

def histogram(df: pd.DataFrame, metric: str, bin_step: int = 5, title: str = "", reference: str = "mean"):
    if metric not in df.columns:
        return alt.Chart(pd.DataFrame())
    clean = df.dropna(subset=[metric])
    hist = alt.Chart(clean).mark_bar().encode(
        x=alt.X(f"{metric}:Q", bin=alt.Bin(step=bin_step), title=metric.replace("_num", " (%)")),
        y=alt.Y("count():Q", title="Count"),
        tooltip=[alt.Tooltip(f"{metric}:Q", title=metric.replace("_num", " (%)")), alt.Tooltip("count():Q", title="Count")]
    ).properties(title=title)
    ref_val = None
    if reference == "median":
        ref_val = float(clean[metric].median()) if not clean.empty else None
    else:
        ref_val = float(clean[metric].mean()) if not clean.empty else None
    if ref_val is not None:
        rule = alt.Chart(pd.DataFrame({"ref": [ref_val]})).mark_rule(color="#d62728").encode(x="ref:Q")
        label = alt.Chart(pd.DataFrame({"ref": [ref_val], "txt": [f"{reference.title()}: {ref_val:.1f}%"]})).mark_text(
            align="left", dx=6, dy=-6, color="#d62728"
        ).encode(x="ref:Q", text="txt:N")
        return hist + rule + label
    return hist

def boxplot(df: pd.DataFrame, dim: str, metric: str, title: str):
    if dim not in df.columns or metric not in df.columns:
        return alt.Chart(pd.DataFrame())
    bp = alt.Chart(df.dropna(subset=[dim, metric])).mark_boxplot(size=22).encode(
        y=alt.Y(f"{dim}:N", title=dim, sort="-x"),
        x=alt.X(f"{metric}:Q", title=metric.replace("_num", " (%)")),
        tooltip=[dim, metric]
    ).properties(title=title)
    return bp

def heatmap(df: pd.DataFrame, row_dim: str, col_dim: str, metric: str, title: str):
    need = [row_dim, col_dim, metric]
    if any(col not in df.columns for col in need):
        return alt.Chart(pd.DataFrame())
    agg = df.dropna(subset=need).groupby([row_dim, col_dim], as_index=False)[metric].mean()
    hm = alt.Chart(agg).mark_rect().encode(
        y=alt.Y(f"{row_dim}:N", title=row_dim, sort="ascending"),
        x=alt.X(f"{col_dim}:N", title=col_dim, sort="ascending"),
        color=alt.Color(f"{metric}:Q", title=metric.replace("_num", " (%)"), scale=alt.Scale(scheme="blues")),
        tooltip=[row_dim, col_dim, alt.Tooltip(metric, format=".1f")]
    ).properties(title=title)
    return hm

def line_trend(df: pd.DataFrame, metric: str, month_col: str = "reporting_month", title: str = "Trend by Month"):
    if metric not in df.columns or month_col not in df.columns:
        return alt.Chart(pd.DataFrame())
    agg = df.dropna(subset=[month_col, metric]).groupby(month_col, as_index=False)[metric].mean()
    ln = alt.Chart(agg).mark_line(point=True).encode(
        x=alt.X(f"{month_col}:N", title="Month"),
        y=alt.Y(f"{metric}:Q", title=metric.replace("_num", " (%)")),
        tooltip=[month_col, alt.Tooltip(metric, format=".1f")]
    ).properties(title=title)
    return ln

# -------------------------------------
# YTD Aggregated Associates Table helper
# -------------------------------------
def ytd_aggregated_table(ytd_df: pd.DataFrame, group_by: str = "Name") -> pd.DataFrame:
    """
    Aggregate Final Score across months for each associate (or domain),
    returning only the requested columns with a dense rank by Final Score.
    """
    if ytd_df is None or ytd_df.empty:
        return pd.DataFrame(columns=["Domain ID","Function","Function Lead","Team Lead","Designation","Name","Final Score","Rank"])
    df = ytd_df.copy()
    if "Final Score_num" not in df.columns and "Final Score" in df.columns:
        s = df["Final Score"].astype(str).str.replace('%','', regex=False).str.replace(',', '.', regex=False)
        df["Final Score_num"] = pd.to_numeric(s, errors='coerce')
    if "Final Score_num" not in df.columns:
        return pd.DataFrame(columns=["Domain ID","Function","Function Lead","Team Lead","Designation","Name","Final Score","Rank"])
    cols_lower = df.columns.str.lower().tolist()
    def has(col): return col.lower() in cols_lower
    group_key = group_by if has(group_by) else ("Domain ID" if has("Domain ID") else None)
    if group_key is None:
        return pd.DataFrame(columns=["Domain ID","Function","Function Lead","Team Lead","Designation","Name","Final Score","Rank"])
    agg_dict = {"Final Score_num": "mean"}
    for c in ["Domain ID","Function","Function Lead","Team Lead","Designation","Name"]:
        if has(c):
            agg_dict[c] = lambda x: x.dropna().iloc[0] if x.dropna().size else None
    grouped = df.groupby(group_key, as_index=False).agg(agg_dict)
    grouped["Final Score"] = grouped["Final Score_num"].round(1)
    rank_series = grouped["Final Score"].rank(method="dense", ascending=False)
    grouped["Rank"] = rank_series.astype("Int64")
    desired_cols = ["Domain ID","Function","Function Lead","Team Lead","Designation","Name","Final Score","Rank"]
    present = [c for c in desired_cols if c in grouped.columns]
    result = grouped[present].copy()
    if "Final Score" in result.columns and "Rank" in result.columns:
        result = result.sort_values(["Rank","Final Score"], ascending=[True, False])
    return result

# -------------------------------------
# Streamlit UI
# -------------------------------------
st.set_page_config(page_title=APP_NAME, layout="wide")
ensure_storage()
ensure_storage_ba()  # existing
ensure_storage_pe()  # NEW
ensure_storage_tl()  # New
ensure_storage_pl()  # New


# Session state
if "authenticated" not in st.session_state:
    st.session_state.authenticated = False
if "role" not in st.session_state:
    st.session_state.role = None
if "username" not in st.session_state:
    st.session_state.username = None
if "hide_cols" not in st.session_state:
    st.session_state.hide_cols = True

def login_block():
    st.sidebar.title("Login")
    u = st.sidebar.text_input("Username")
    p = st.sidebar.text_input("Password", type="password")
    if st.sidebar.button("Sign In"):
        ok, role = authenticate(u, p)
        if ok:
            st.session_state.authenticated = True
            st.session_state.role = role
            st.session_state.username = u
            st.sidebar.success(f"Logged in as {USERS[u]['display_name']} ({role})")
        else:
            st.sidebar.error("Invalid credentials")

if not st.session_state.authenticated:
    login_block()
    st.stop()

st.sidebar.title(APP_NAME)
# ✅ Global toggle (applies to ALL pages & downloads)
st.sidebar.checkbox("Hide “Unnamed” & empty columns", value=st.session_state.hide_cols, key="hide_cols")

pages = [
    "Associates Scorecard (Monthly/YTD metrics)",
    "BA Scorecard (Monthly/YTD metrics)",
    "PE Scorecard (Monthly/YTD metrics)",  
    "TL Scorecard (Monthly/YTD metrics)", 
    "PL Scorecard (Monthly/YTD metrics)", # NEW
    "History"
]

if st.session_state.role == "admin":
    pages.append("Upload & Admin")
page = st.sidebar.radio("Navigate", pages)

# -------------------------------------
# Unified Page: Associates Scorecard (Monthly/YTD metrics)
# -------------------------------------
if page == "Associates Scorecard (Monthly/YTD metrics)":
    st.header("📊 Associates Scorecard (Monthly/YTD metrics)")
    mode = st.radio("View mode", ["Monthly", "YTD"], index=0, horizontal=True)

    # ===== Monthly =====
    if mode == "Monthly":

        ### replacing the below line 1 with code snippet with comment as "Use cached reads to assemble latest data"
        ### latest_row, latest_id, latest_data = get_latest_monthly_data()

        # Use cached reads to assemble latest data
        h = load_history_cached()
        active_mask = _coerce_active_bool(h.get("active", pd.Series([], dtype="object")))
        active = h[active_mask]
        if active.empty:
            latest_row, latest_id, latest_data = None, None, None
        else:
            latest_row = active.sort_values("upload_dt", ascending=False).iloc[0]
            latest_id = latest_row["id"]
            combined_df = load_combined_cached()
            latest_data = combined_df[combined_df["Attachment ID"] == latest_id]

        if latest_data is None or latest_data.empty:
            st.warning("No active file available.")
        else:
            latest_data = clean_dataframe_for_display(latest_data, st.session_state.hide_cols)
            latest_data = add_numeric_cached(latest_data)


            def render_shared_filters(df, label="Filters (Monthly)"):
                month_str = _to_month_str_series(df)
                month_options = sorted([m for m in month_str.dropna().unique() if m and str(m).strip() != ""])
                with st.expander(label, expanded=True):
                    c1, c2, c3, c4, c5, c6 = st.columns(6)
                    domain_options = sorted(df["Domain ID"].dropna().astype(str).unique()) if "Domain ID" in df.columns else []
                    func_options = sorted(df["Function"].dropna().astype(str).unique()) if "Function" in df.columns else []
                    flead_options = sorted(df["Function Lead"].dropna().astype(str).unique()) if "Function Lead" in df.columns else []
                    tlead_options = sorted(df["Team Lead"].dropna().astype(str).unique()) if "Team Lead" in df.columns else []
                    d_ids = c1.multiselect("Domain ID (select one to comment)", domain_options)
                    funcs = c2.multiselect("Function", func_options)
                    f_leads = c3.multiselect("Function Lead", flead_options)
                    t_leads = c4.multiselect("Team Lead", tlead_options)
                    # 👉 New: Final score band next to Team Lead
                    final_score_band = c5.selectbox("Final score",options=["All", ">= 100", "Between 90 and 99.99", "< 90"],index=0 )
                    months = c6.multiselect("Month (YYYY-MM)", month_options)

                    # Feedback entry when exactly one Domain ID and one Month
                    if d_ids and len(d_ids) == 1 and months and len(months) == 1:
                        sel_domain = str(d_ids[0])
                        sel_month = str(months[0])
                        st.markdown("**Monthly feedback/feedforward** (max 500 characters)")
                        fb = load_feedback()
                        existing = fb[(fb["Domain ID"].astype(str) == sel_domain) & (fb["Month"].astype(str) == sel_month)]
                        existing_text = existing.iloc[0]["Feedback"] if not existing.empty else ""
                        comment = st.text_area(
                            "Enter feedback for the selected Domain ID & Month",
                            value=str(existing_text),
                            max_chars=MAX_FEEDBACK_CHARS,
                            height=120
                        )
                        password = st.text_input("Enter Team Lead password to confirm", type="password")
                        if st.button("Submit Comment"):
                            if password == FEEDBACK_PASSWORD:
                                sel_df = df[df["Domain ID"].astype(str) == sel_domain]
                                if not sel_df.empty:
                                    name_val = sel_df.iloc[0]["Name"] if "Name" in sel_df.columns else sel_domain
                                    tl_val = sel_df.iloc[0]["Team Lead"] if "Team Lead" in sel_df.columns else ""
                                    if comment and comment.strip():
                                        upsert_feedback(sel_domain, str(name_val), sel_month, str(tl_val),
                                                        comment.strip(), st.session_state.username or "user")
                                        st.success("Feedback saved.")
                                    else:
                                        st.warning("Please enter a feedback comment before submitting.")
                                else:
                                    st.error("Selected Domain ID not found in current dataset.")
                            else:
                                st.error("Incorrect password. Feedback not saved.")
                return d_ids, funcs, f_leads, t_leads, months, final_score_band, None

            # Below 2 lines of code are the older ones.
            # d_ids, funcs, f_leads, t_leads, months, _ = render_shared_filters(latest_data)
            # filtered = filter_combined(latest_data, d_ids, funcs, f_leads, t_leads, months)

            
            d_ids, funcs, f_leads, t_leads, months, fs_band, _ = render_shared_filters(latest_data)
            filtered = filter_combined(latest_data, d_ids, funcs, f_leads, t_leads, months)
            # 👉 Apply Final score band filter
            filtered = apply_final_score_band_filter(filtered, fs_band)
            filtered = clean_dataframe_for_display(filtered, st.session_state.hide_cols)

            c1,c2,c3,c4 = st.columns(4)
            c1.metric("Active Month", latest_row["reporting_month"])
            c2.metric("Rows (after filters)", len(filtered))
            c3.metric("Uploader", latest_row["uploader"])
            c4.metric("Last Upload", latest_row["upload_dt"])

            st.subheader("Associate Monthly Metrics")
            active_month = str(latest_row["reporting_month"])
            if months and len(months) == 1:
                active_month = str(months[0])

            mon_metrics = monthly_metrics_table(
                filtered,
                report_month=active_month,
                group_by="Domain ID"
            )
            st.caption(f"Showing {len(mon_metrics)} monthly rows (from filtered view)")
            
            # 👉 Styled display (Associates Monthly only)
            mon_metrics_styled = style_associates_metrics_df(mon_metrics)
            # Before (less reliable for Styler colors)
            # st.dataframe(mon_metrics, height=420)

            st.table(mon_metrics_styled)
            
            # 👉 Download with colors applied in Excel (Associates Monthly only)
            if exceeds_excel_limits(mon_metrics):
                st.caption("Note: Monthly metrics are too wide for Excel; download provided as CSV.")
            st.download_button(
                "⬇️ Download Monthly Metrics",
                make_excel_bytes_associates_monthly_metrics(mon_metrics, st.session_state.hide_cols),
                file_name=f"{EXPORT_PREFIX}monthly_associates_metrics_{active_month}.xlsx"
            )


            # Simple charts
            if "Function" in filtered.columns and "Final Score_num" in filtered.columns:
                final_func = (
                    filtered.dropna(subset=["Final Score_num", "Function"])
                            .groupby("Function", as_index=False)["Final Score_num"].mean()
                            .rename(columns={"Final Score_num":"Avg Final Score (%)"})
                )
                sel_func = alt.selection_multi(fields=["Function"], bind="legend")
                chart_a = alt.Chart(final_func).mark_bar().encode(
                    x=alt.X("Avg Final Score (%):Q", title="Avg Final Score (%)"),
                    y=alt.Y("Function:N", sort="-x"),
                    color=alt.Color("Function:N", legend=alt.Legend(title="Click legend to filter")),
                    tooltip=["Function","Avg Final Score (%)"]
                ).add_selection(sel_func).properties(title="Avg Final Score by Function (Monthly)")
                st.altair_chart(chart_a, use_container_width=True)
            else:
                st.info("Final Score or Function column not found—'Avg Final Score by Function' chart skipped.")

            if "Final Score_num" in filtered.columns:
                st.altair_chart(
                    alt.Chart(filtered.dropna(subset=["Final Score_num"]))
                      .mark_bar()
                      .encode(
                          x=alt.X("Final Score_num:Q", bin=alt.Bin(step=5), title="Final Score (%)"),
                          y=alt.Y("count():Q", title="Count"),
                          tooltip=[alt.Tooltip("Final Score_num:Q", title="Final Score (%)"), alt.Tooltip("count():Q", title="Count")]
                      ).properties(title="Final Score Distribution (Monthly, 5% bins)"),
                    use_container_width=True
                )

            enable_altair_theme()
            with st.expander("🎨 Advanced Visualizations (Monthly)", expanded=False):
                if filtered is None or filtered.empty:
                    st.info("No data under current filters for advanced visuals.")
                else:
                    cset1, cset2, cset3, cset4 = st.columns([2,2,2,2])
                    metric_options = get_numeric_metric_options(filtered)
                    default_metric_list = metric_options if metric_options else ["Final Score_num"]
                    default_index = default_metric_list.index("Final Score_num") if "Final Score_num" in default_metric_list else 0
                    sel_metric = cset1.selectbox("Metric (numeric %)", options=default_metric_list, index=default_index)
                    agg_method = cset2.radio("Aggregation", ["mean", "median"], index=0)
                    dim_candidates = [c for c in ["Function","Team Lead","Function Lead","Domain ID"] if c in filtered.columns]
                    dim = cset3.selectbox("Group by", options=dim_candidates if dim_candidates else ["Function"], index=0)
                    palette = cset4.selectbox("Palette", options=list(PALETTES.keys()), index=list(PALETTES.keys()).index("Tableau10"))
                    cN1, cN2, cN3 = st.columns([1,1,1])
                    top_n = cN1.slider("Top N", min_value=5, max_value=50, value=15, step=5)
                    ascending = cN2.checkbox("Show lowest first", value=False)
                    show_labels = cN3.checkbox("Bar labels", value=True)

                    agg_df = aggregate_df(filtered, dim=dim, metric=sel_metric, method=agg_method)
                    agg_df = add_rank_and_topN(agg_df, dim=dim, metric=sel_metric, top_n=top_n, ascending=ascending)
                    st.altair_chart(
                        bar_chart(agg_df, dim=dim, metric=sel_metric,
                                  title=f"{agg_method.title()} {sel_metric.replace('_num',' (%)')} by {dim} (Monthly)",
                                  palette=palette, show_labels=show_labels),
                        use_container_width=True
                    )

                    cH1, cH2 = st.columns([1,1])
                    bin_step = cH1.slider("Histogram bin step (percentage points)", 1, 20, 5, 1)
                    ref = cH2.radio("Reference line", ["mean", "median"], index=0)
                    st.altair_chart(
                        histogram(filtered, metric=sel_metric, bin_step=bin_step,
                                  title=f"Distribution of {sel_metric.replace('_num',' (%)')} (Monthly)",
                                  reference=ref),
                        use_container_width=True
                    )

                    st.altair_chart(
                        boxplot(filtered, dim=dim, metric=sel_metric, title=f"Distribution by {dim} (Monthly)"),
                        use_container_width=True
                    )

                    try:
                        ### history = load_history() below is the new  line of code
                        history = load_history_cached()
                        active_mask_trend = _coerce_active_bool(history.get("active", pd.Series([], dtype="object")))
                        active_ids = history[active_mask_trend][["id","reporting_month"]].rename(columns={"id":"Attachment ID"})
                        ### monthly_all = load_combined() below is the new  line of code
                        monthly_all = load_combined_cached()
                        monthly_active = monthly_all.merge(active_ids, on="Attachment ID", how="inner")
                        monthly_active["reporting_month"] = monthly_active["reporting_month"].astype(str)
                        monthly_active = add_numeric_percent_columns(monthly_active)
                        st.altair_chart(
                            line_trend(monthly_active, metric=sel_metric, month_col="reporting_month",
                                       title=f"Trend by Month (active attachments): {sel_metric.replace('_num',' (%)')}"),
                            use_container_width=True
                        )
                    except Exception:
                        st.caption("Trend by month unavailable under current data.")

            st.subheader("Filtered Table (latest active file)")
            st.caption(f"Showing {len(filtered)} of {len(latest_data)} rows")
            st.dataframe(filtered if not filtered.empty else pd.DataFrame(), height=480)
            if exceeds_excel_limits(filtered):
                st.caption("Note: Filtered result is too wide for Excel; download provided as CSV.")
            st.download_button(
                "⬇️ Download filtered (Monthly)",
                make_excel_bytes_from_df(filtered, st.session_state.hide_cols),
                file_name=f"{EXPORT_PREFIX}monthly_scorecard_filtered.xlsx"
            )

            if st.session_state.role == "admin":
                st.subheader("🛠 Admin — Edit Latest Active Data")
                st.caption("Edit values directly. Saving replaces the data for the latest active attachment in combined storage (not the original Excel file).")
                editable = st.data_editor(latest_data.copy(), num_rows="dynamic", use_container_width=True)
                if st.button("Save Admin Changes", type="primary"):
                    try:
                        edited = editable.copy()
                        edited = convert_percentage_columns(edited)
                        edited = add_numeric_percent_columns(edited)
                        edited["Attachment ID"] = latest_id
                        combined_df = load_combined()
                        combined_df = combined_df[combined_df["Attachment ID"] != latest_id]
                        combined_df = pd.concat([combined_df, edited], ignore_index=True)
                        save_combined(combined_df)
                        log_audit("Admin Save Edit",
                                  latest_id,
                                  latest_row["filename"] if "filename" in latest_row else "",
                                  st.session_state.username or "admin")
                        st.success("Admin changes saved to combined storage.")
                        invalidate_data_caches()  # ensure next UI run fetches fresh files
                    except Exception as e:
                        st.error(f"Failed to save admin changes: {e}")


    # ===== YTD =====
    else:
        ### history = load_history() below is the new  line of code
        history = load_history_cached()
        active_mask = _coerce_active_bool(history.get("active", pd.Series([], dtype="object")))
        active_ids = history[active_mask][["id","reporting_month"]].rename(columns={"id":"Attachment ID"})
        ### monthly_all = load_combined() below is the new  line of code
        monthly_all = load_combined_cached()
        ytd = monthly_all.merge(active_ids, on="Attachment ID", how="inner")
        if ytd.empty:
            st.warning("No YTD data.")
            st.stop()
        ytd = clean_dataframe_for_display(ytd, st.session_state.hide_cols)
        ytd = add_numeric_cached(ytd)

        def render_ytd_filters(df, label="Filters (YTD)"):
            month_str = _to_month_str_series(df)
            month_options = sorted([m for m in month_str.dropna().unique() if m and str(m).strip() != ""]) if not month_str.empty else []
            def _fy_label(ym: str) -> str:
                try:
                    y, m = ym.split('-'); y = int(y); m = int(m)
                except Exception:
                    return ""
                if m >= 4: fy_start, fy_end = y, y+1
                else: fy_start, fy_end = y-1, y
                return f"FY{fy_start}-{str(fy_end)[-2:]}"
            with st.expander(label, expanded=True):
                c1, c2, c3, c4, c5, c6 = st.columns(6)
                domain_options = sorted(df["Domain ID"].dropna().astype(str).unique()) if "Domain ID" in df.columns else []
                func_options = sorted(df["Function"].dropna().astype(str).unique()) if "Function" in df.columns else []
                flead_options = sorted(df["Function Lead"].dropna().astype(str).unique()) if "Function Lead" in df.columns else []
                tlead_options = sorted(df["Team Lead"].dropna().astype(str).unique()) if "Team Lead" in df.columns else []
                d_ids = c1.multiselect("Domain ID", domain_options)
                funcs = c2.multiselect("Function", func_options)
                f_leads = c3.multiselect("Function Lead", flead_options)
                t_leads = c4.multiselect("Team Lead", tlead_options)

                # 👉 New: Final score band next to Team Lead (same options as Monthly)
                final_score_band = c5.selectbox("Final score",options=["All", ">= 100", "Between 90 and 99.99", "< 90"],index=0)


                use_fy = c6.checkbox("Use Fiscal Year (Apr–Mar)", value=False)
                if use_fy and month_options:
                    fy_options = sorted({ _fy_label(m) for m in month_options if _fy_label(m) })
                    sel_fy = c6.selectbox("Fiscal Year", options=fy_options, index=0 if fy_options else 0)
                    months = [m for m in month_options if _fy_label(m) == sel_fy]
                    st.caption(f"Months auto-selected for {sel_fy}: {', '.join(months)}")
                else:
                    months = c6.multiselect("Month (YYYY-MM)", month_options)
                return d_ids, funcs, f_leads, t_leads, months, final_score_band
        
        ### Below are the 2 older lines
        ### d_ids, funcs, f_leads, t_leads, months, search = render_ytd_filters(ytd)
        ### ytd_filtered = filter_combined(ytd, d_ids, funcs, f_leads, t_leads, months)

        #Now:
        d_ids, funcs, f_leads, t_leads, months, fs_band = render_ytd_filters(ytd)
        search = st.text_input("🔎 Search across all columns (YTD)")  # keep existing search UI
        ytd_filtered = filter_combined(ytd, d_ids, funcs, f_leads, t_leads, months)
        # 👉 Apply Final score band (same as Monthly approach)
        ytd_filtered = apply_final_score_band_filter(ytd_filtered, fs_band)
        ytd_filtered = apply_search(ytd_filtered, search)
        ytd_filtered = clean_dataframe_for_display(ytd_filtered, st.session_state.hide_cols)

        c1, c2, c3 = st.columns(3)
        c1.metric("YTD Rows (after filters)", len(ytd_filtered))
        c2.metric("Distinct Domains", ytd_filtered["Domain ID"].nunique() if "Domain ID" in ytd_filtered.columns else 0)
        c3.metric("Distinct Functions", ytd_filtered["Function"].nunique() if "Function" in ytd_filtered.columns else 0)

        st.subheader("YTD Aggregated Associates Table")
        agg_options = [opt for opt in ["Name","Domain ID"] if opt in ytd_filtered.columns]
        agg_by = st.selectbox("Aggregate by", options=agg_options, index=0 if "Name" in agg_options else 0)
        ytd_assoc_agg = ytd_aggregated_table(ytd_filtered, group_by=agg_by)
        st.caption(f"Showing {len(ytd_assoc_agg)} aggregated rows")
        st.dataframe(ytd_assoc_agg, height=420)
     
        # 👉 Add color-coded view (same scheme as Monthly)
        ytd_assoc_agg_styled = style_associates_metrics_df(ytd_assoc_agg)
        st.caption("Color-coded view (Final Score)")
        st.table(ytd_assoc_agg_styled)

        if exceeds_excel_limits(ytd_assoc_agg):
            st.caption("Note: Aggregated result is too wide for Excel; download provided as CSV.")
        
        st.download_button(
            "⬇️ Download aggregated (YTD Associates Table)",
            make_excel_bytes_from_df(ytd_assoc_agg, st.session_state.hide_cols),
            file_name=f"{EXPORT_PREFIX}ytd_associates_aggregated_{agg_by.lower().replace(' ', '_')}.xlsx"
        )

        # 👉 New: color-coded Excel download for YTD
        st.download_button(
            "⬇️ Download aggregated (YTD Associates Table – colored)",
            make_excel_bytes_associates_ytd_aggregated(ytd_assoc_agg, st.session_state.hide_cols),
            file_name=f"{EXPORT_PREFIX}ytd_associates_aggregated_{agg_by.lower().replace(' ', '_')}_colored.xlsx"
        )


        enable_altair_theme()
        with st.expander("🎨 Advanced Visualizations (YTD)", expanded=False):
            if ytd_filtered is None or ytd_filtered.empty:
                st.info("No YTD data under current filters for advanced visuals.")
            else:
                cset1, cset2, cset3, cset4 = st.columns([2,2,2,2])
                metric_options_ytd = get_numeric_metric_options(ytd_filtered)
                default_metric_list_ytd = metric_options_ytd if metric_options_ytd else ["Final Score_num"]
                default_index_ytd = default_metric_list_ytd.index("Final Score_num") if "Final Score_num" in default_metric_list_ytd else 0
                sel_metric_ytd = cset1.selectbox("Metric (numeric %)", options=default_metric_list_ytd, index=default_index_ytd)
                agg_method_ytd = cset2.radio("Aggregation", ["mean", "median"], index=0)
                dim_candidates_ytd = [c for c in ["Function","Team Lead","Function Lead","Domain ID"] if c in ytd_filtered.columns]
                dim_ytd = cset3.selectbox("Group by", options=dim_candidates_ytd if dim_candidates_ytd else ["Function"], index=0)
                palette_ytd = cset4.selectbox("Palette", options=list(PALETTES.keys()), index=list(PALETTES.keys()).index("Tableau10"))
                cN1, cN2, cN3 = st.columns([1,1,1])
                top_n_ytd = cN1.slider("Top N", min_value=5, max_value=50, value=15, step=5)
                ascending_ytd = cN2.checkbox("Show lowest first", value=False)
                show_labels_ytd = cN3.checkbox("Bar labels", value=True)

                agg_ytd = aggregate_df(ytd_filtered, dim=dim_ytd, metric=sel_metric_ytd, method=agg_method_ytd)
                agg_ytd = add_rank_and_topN(agg_ytd, dim=dim_ytd, metric=sel_metric_ytd, top_n=top_n_ytd, ascending=ascending_ytd)
                st.altair_chart(
                    bar_chart(agg_ytd, dim=dim_ytd, metric=sel_metric_ytd,
                              title=f"{agg_method_ytd.title()} {sel_metric_ytd.replace('_num',' (%)')} by {dim_ytd} (YTD)",
                              palette=palette_ytd, show_labels=show_labels_ytd),
                    use_container_width=True
                )

                cH1, cH2 = st.columns([1,1])
                bin_step_ytd = cH1.slider("Histogram bin step (percentage points)", 1, 20, 5, 1)
                ref_ytd = cH2.radio("Reference line", ["mean", "median"], index=0)
                st.altair_chart(
                    histogram(ytd_filtered, metric=sel_metric_ytd, bin_step=bin_step_ytd,
                              title=f"Distribution of {sel_metric_ytd.replace('_num',' (%)')} (YTD)",
                              reference=ref_ytd),
                    use_container_width=True
                )

                st.altair_chart(
                    boxplot(ytd_filtered, dim=dim_ytd, metric=sel_metric_ytd, title=f"Distribution by {dim_ytd} (YTD)"),
                    use_container_width=True
                )

                if "Function" in ytd_filtered.columns and "Team Lead" in ytd_filtered.columns:
                    st.altair_chart(
                        heatmap(ytd_filtered, row_dim="Function", col_dim="Team Lead", metric=sel_metric_ytd,
                                title=f"Heatmap: {sel_metric_ytd.replace('_num',' (%)')} (Function x Team Lead) - YTD"),
                        use_container_width=True
                    )
                else:
                    ytd_norm = ytd_filtered.copy()
                    ytd_norm["Month_norm"] = _to_month_str_series(ytd_norm)
                    if "Function" in ytd_norm.columns and "Month_norm" in ytd_norm.columns:
                        st.altair_chart(
                            heatmap(ytd_norm, row_dim="Function", col_dim="Month_norm", metric=sel_metric_ytd,
                                    title=f"Heatmap: {sel_metric_ytd.replace('_num',' (%)')} (Function x Month) - YTD"),
                            use_container_width=True
                        )

        st.subheader("Filtered YTD Table")
        st.caption(f"Showing {len(ytd_filtered)} of {len(ytd)} rows")
        st.dataframe(ytd_filtered, height=480)
        if exceeds_excel_limits(ytd_filtered):
            st.caption("Note: Filtered result is too wide for Excel; download provided as CSV.")
        st.download_button(
            "⬇️ Download filtered (YTD)",
            make_excel_bytes_from_df(ytd_filtered, st.session_state.hide_cols),
            file_name=f"{EXPORT_PREFIX}ytd_dashboard_filtered.xlsx"
        )

# -------------------------------------
# New Page: BA Scorecard (Monthly/YTD metrics)
# -------------------------------------
elif page == "BA Scorecard (Monthly/YTD metrics)":
    st.header("📊 BA Scorecard (Monthly/YTD metrics)")
    mode = st.radio("View mode", ["Monthly", "YTD"], index=0, horizontal=True)

    if mode == "Monthly":
        latest_row, latest_id, latest_data = ba_get_latest_monthly_data()
        if latest_data is None or latest_data.empty:
            st.warning("No active BA file available.")
        else:
            latest_data = clean_dataframe_for_display(latest_data, st.session_state.hide_cols)
            latest_data = add_numeric_cached(latest_data)

            def render_shared_filters_ba(df, label="Filters (Monthly)"):
                month_str = _to_month_str_series(df)
                month_options = sorted([m for m in month_str.dropna().unique() if m and str(m).strip() != ""])
                with st.expander(label, expanded=True):
                    c1, c2, c3, c4, c5 = st.columns(5)
                    domain_options = sorted(df["Domain ID"].dropna().astype(str).unique()) if "Domain ID" in df.columns else []
                    func_options = sorted(df["Function"].dropna().astype(str).unique()) if "Function" in df.columns else []
                    flead_options = sorted(df["Function Lead"].dropna().astype(str).unique()) if "Function Lead" in df.columns else []
                    tlead_options = sorted(df["Team Lead"].dropna().astype(str).unique()) if "Team Lead" in df.columns else []
                    d_ids = c1.multiselect("Domain ID (select one to comment)", domain_options)
                    funcs = c2.multiselect("Function", func_options)
                    f_leads = c3.multiselect("Function Lead", flead_options)
                    t_leads = c4.multiselect("Team Lead", tlead_options)
                    months = c5.multiselect("Month (YYYY-MM)", month_options)

                    if d_ids and len(d_ids) == 1 and months and len(months) == 1:
                        sel_domain = str(d_ids[0])
                        sel_month = str(months[0])
                        st.markdown("**Monthly feedback/feedforward** (max 500 characters)")
                        fb = ba_load_feedback()
                        existing = fb[(fb["Domain ID"].astype(str) == sel_domain) & (fb["Month"].astype(str) == sel_month)]
                        existing_text = existing.iloc[0]["Feedback"] if not existing.empty else ""
                        comment = st.text_area(
                            "Enter feedback for the selected Domain ID & Month",
                            value=str(existing_text),
                            max_chars=MAX_FEEDBACK_CHARS,
                            height=120
                        )
                        password = st.text_input("Enter Team Lead password to confirm", type="password")
                        if st.button("Submit Comment"):
                            if password == FEEDBACK_PASSWORD:
                                sel_df = df[df["Domain ID"].astype(str) == sel_domain]
                                if not sel_df.empty:
                                    name_val = sel_df.iloc[0]["Name"] if "Name" in sel_df.columns else sel_domain
                                    tl_val = sel_df.iloc[0]["Team Lead"] if "Team Lead" in sel_df.columns else ""
                                    if comment and comment.strip():
                                        ba_upsert_feedback(sel_domain, str(name_val), sel_month, str(tl_val),
                                                           comment.strip(), st.session_state.username or "user")
                                        st.success("Feedback saved.")
                                    else:
                                        st.warning("Please enter a feedback comment before submitting.")
                                else:
                                    st.error("Selected Domain ID not found in current dataset.")
                            else:
                                st.error("Incorrect password. Feedback not saved.")
                return d_ids, funcs, f_leads, t_leads, months, None

            d_ids, funcs, f_leads, t_leads, months, _ = render_shared_filters_ba(latest_data)
            filtered = filter_combined(latest_data, d_ids, funcs, f_leads, t_leads, months)
            filtered = clean_dataframe_for_display(filtered, st.session_state.hide_cols)

            c1,c2,c3,c4 = st.columns(4)
            c1.metric("Active Month", latest_row["reporting_month"])
            c2.metric("Rows (after filters)", len(filtered))
            c3.metric("Uploader", latest_row["uploader"])
            c4.metric("Last Upload", latest_row["upload_dt"])

            st.subheader("BA Monthly Metrics")
            active_month = str(latest_row["reporting_month"])
            if months and len(months) == 1:
                active_month = str(months[0])

            mon_metrics = monthly_metrics_table_ba(filtered, report_month=active_month, group_by="Domain ID")
            st.caption(f"Showing {len(mon_metrics)} monthly rows (from filtered view)")
            st.dataframe(mon_metrics, height=420)

            if exceeds_excel_limits(mon_metrics):
                st.caption("Note: Monthly metrics are too wide for Excel; download provided as CSV.")
            st.download_button(
                "⬇️ Download Monthly Metrics",
                make_excel_bytes_from_df(mon_metrics, st.session_state.hide_cols),
                file_name=f"{BA_EXPORT_PREFIX}monthly_ba_metrics_{active_month}.xlsx"
            )

            if "Function" in filtered.columns and "Final Score_num" in filtered.columns:
                final_func = (
                    filtered.dropna(subset=["Final Score_num", "Function"])
                            .groupby("Function", as_index=False)["Final Score_num"].mean()
                            .rename(columns={"Final Score_num":"Avg Final Score (%)"})
                )
                sel_func = alt.selection_multi(fields=["Function"], bind="legend")
                chart_a = alt.Chart(final_func).mark_bar().encode(
                    x=alt.X("Avg Final Score (%):Q", title="Avg Final Score (%)"),
                    y=alt.Y("Function:N", sort="-x"),
                    color=alt.Color("Function:N", legend=alt.Legend(title="Click legend to filter")),
                    tooltip=["Function","Avg Final Score (%)"]
                ).add_selection(sel_func).properties(title="Avg Final Score by Function (Monthly)")
                st.altair_chart(chart_a, use_container_width=True)
            else:
                st.info("Final Score or Function column not found—'Avg Final Score by Function' chart skipped.")

            if "Final Score_num" in filtered.columns:
                st.altair_chart(
                    alt.Chart(filtered.dropna(subset=["Final Score_num"]))
                      .mark_bar()
                      .encode(
                          x=alt.X("Final Score_num:Q", bin=alt.Bin(step=5), title="Final Score (%)"),
                          y=alt.Y("count():Q", title="Count"),
                          tooltip=[alt.Tooltip("Final Score_num:Q", title="Final Score (%)"), alt.Tooltip("count():Q", title="Count")]
                      ).properties(title="Final Score Distribution (Monthly, 5% bins)"),
                    use_container_width=True
                )

            enable_altair_theme()
            with st.expander("🎨 Advanced Visualizations (Monthly)", expanded=False):
                if filtered is None or filtered.empty:
                    st.info("No data under current filters for advanced visuals.")
                else:
                    cset1, cset2, cset3, cset4 = st.columns([2,2,2,2])
                    metric_options = get_numeric_metric_options(filtered)
                    default_metric_list = metric_options if metric_options else ["Final Score_num"]
                    default_index = default_metric_list.index("Final Score_num") if "Final Score_num" in default_metric_list else 0
                    sel_metric = cset1.selectbox("Metric (numeric %)", options=default_metric_list, index=default_index)
                    agg_method = cset2.radio("Aggregation", ["mean", "median"], index=0)
                    dim_candidates = [c for c in ["Function","Team Lead","Function Lead","Domain ID"] if c in filtered.columns]
                    dim = cset3.selectbox("Group by", options=dim_candidates if dim_candidates else ["Function"], index=0)
                    palette = cset4.selectbox("Palette", options=list(PALETTES.keys()), index=list(PALETTES.keys()).index("Tableau10"))
                    cN1, cN2, cN3 = st.columns([1,1,1])
                    top_n = cN1.slider("Top N", min_value=5, max_value=50, value=15, step=5)
                    ascending = cN2.checkbox("Show lowest first", value=False)
                    show_labels = cN3.checkbox("Bar labels", value=True)

                    agg_df = aggregate_df(filtered, dim=dim, metric=sel_metric, method=agg_method)
                    agg_df = add_rank_and_topN(agg_df, dim=dim, metric=sel_metric, top_n=top_n, ascending=ascending)
                    st.altair_chart(
                        bar_chart(agg_df, dim=dim, metric=sel_metric,
                                  title=f"{agg_method.title()} {sel_metric.replace('_num',' (%)')} by {dim} (Monthly)",
                                  palette=palette, show_labels=show_labels),
                        use_container_width=True
                    )

                    cH1, cH2 = st.columns([1,1])
                    bin_step = cH1.slider("Histogram bin step (percentage points)", 1, 20, 5, 1)
                    ref = cH2.radio("Reference line", ["mean", "median"], index=0)
                    st.altair_chart(
                        histogram(filtered, metric=sel_metric, bin_step=bin_step,
                                  title=f"Distribution of {sel_metric.replace('_num',' (%)')} (Monthly)",
                                  reference=ref),
                        use_container_width=True
                    )

                    st.altair_chart(
                        boxplot(filtered, dim=dim, metric=sel_metric, title=f"Distribution by {dim} (Monthly)"),
                        use_container_width=True
                    )

                    try:
                        history = ba_load_history_cached()
                        active_mask_trend = _coerce_active_bool(history.get("active", pd.Series([], dtype="object")))
                        active_ids = history[active_mask_trend][["id","reporting_month"]].rename(columns={"id":"Attachment ID"})
                        monthly_all = ba_load_combined_cached()
                        monthly_active = monthly_all.merge(active_ids, on="Attachment ID", how="inner")
                        monthly_active["reporting_month"] = monthly_active["reporting_month"].astype(str)
                        monthly_active = add_numeric_percent_columns(monthly_active)
                        st.altair_chart(
                            line_trend(monthly_active, metric=sel_metric, month_col="reporting_month",
                                       title=f"Trend by Month (active attachments): {sel_metric.replace('_num',' (%)')}"),
                            use_container_width=True
                        )
                    except Exception:
                        st.caption("Trend by month unavailable under current data.")

            st.subheader("Filtered Table (latest active BA file)")
            st.caption(f"Showing {len(filtered)} of {len(latest_data)} rows")
            st.dataframe(filtered if not filtered.empty else pd.DataFrame(), height=480)
            if exceeds_excel_limits(filtered):
                st.caption("Note: Filtered result is too wide for Excel; download provided as CSV.")
            st.download_button(
                "⬇️ Download filtered (Monthly)",
                make_excel_bytes_from_df(filtered, st.session_state.hide_cols),
                file_name=f"{BA_EXPORT_PREFIX}monthly_scorecard_filtered.xlsx"
            )

            if st.session_state.role == "admin":
                st.subheader("🛠 Admin — Edit Latest Active BA Data")
                st.caption("Edit values directly. Saving replaces the data for the latest active BA attachment in combined storage (not the original Excel file).")
                editable = st.data_editor(latest_data.copy(), num_rows="dynamic", use_container_width=True)
                if st.button("Save Admin Changes (BA)", type="primary"):
                    try:
                        edited = editable.copy()
                        edited = convert_percentage_columns(edited)
                        edited = add_numeric_percent_columns(edited)
                        edited["Attachment ID"] = latest_id
                        combined_df = ba_load_combined_cached()
                        combined_df = combined_df[combined_df["Attachment ID"] != latest_id]
                        combined_df = pd.concat([combined_df, edited], ignore_index=True)
                        ba_save_combined(combined_df)
                        ba_log_audit("Admin Save Edit (BA)",
                                     latest_id,
                                     latest_row["filename"] if "filename" in latest_row else "",
                                     st.session_state.username or "admin")
                        st.success("Admin changes saved to BA combined storage.")
                        invalidate_data_caches()  # ensure next UI run fetches fresh files
                    except Exception as e:
                        st.error(f"Failed to save admin changes: {e}")

    else:
        history = ba_load_history_cached()
        active_mask = _coerce_active_bool(history.get("active", pd.Series([], dtype="object")))
        active_ids = history[active_mask][["id","reporting_month"]].rename(columns={"id":"Attachment ID"})
        monthly_all = ba_load_combined_cached()
        ytd = monthly_all.merge(active_ids, on="Attachment ID", how="inner")
        if ytd.empty:
            st.warning("No BA YTD data.")
            st.stop()
        ytd = clean_dataframe_for_display(ytd, st.session_state.hide_cols)
        ytd = add_numeric_cached(ytd)

        def render_ytd_filters_ba(df, label="Filters (YTD)"):
            month_str = _to_month_str_series(df)
            month_options = sorted([m for m in month_str.dropna().unique() if m and str(m).strip() != ""]) if not month_str.empty else []
            def _fy_label(ym: str) -> str:
                try:
                    y, m = ym.split('-'); y = int(y); m = int(m)
                except Exception:
                    return ""
                if m >= 4: fy_start, fy_end = y, y+1
                else: fy_start, fy_end = y-1, y
                return f"FY{fy_start}-{str(fy_end)[-2:]}"
            with st.expander(label, expanded=True):
                c1, c2, c3, c4, c5 = st.columns(5)
                domain_options = sorted(df["Domain ID"].dropna().astype(str).unique()) if "Domain ID" in df.columns else []
                func_options = sorted(df["Function"].dropna().astype(str).unique()) if "Function" in df.columns else []
                flead_options = sorted(df["Function Lead"].dropna().astype(str).unique()) if "Function Lead" in df.columns else []
                tlead_options = sorted(df["Team Lead"].dropna().astype(str).unique()) if "Team Lead" in df.columns else []
                d_ids = c1.multiselect("Domain ID", domain_options)
                funcs = c2.multiselect("Function", func_options)
                f_leads = c3.multiselect("Function Lead", flead_options)
                t_leads = c4.multiselect("Team Lead", tlead_options)
                use_fy = c5.checkbox("Use Fiscal Year (Apr–Mar)", value=False)
                if use_fy and month_options:
                    fy_options = sorted({ _fy_label(m) for m in month_options if _fy_label(m) })
                    sel_fy = c5.selectbox("Fiscal Year", options=fy_options, index=0 if fy_options else 0)
                    months = [m for m in month_options if _fy_label(m) == sel_fy]
                    st.caption(f"Months auto-selected for {sel_fy}: {', '.join(months)}")
                else:
                    months = c5.multiselect("Month (YYYY-MM)", month_options)
                search = st.text_input("🔎 Search across all columns (YTD)")
                return d_ids, funcs, f_leads, t_leads, months, search

        d_ids, funcs, f_leads, t_leads, months, search = render_ytd_filters_ba(ytd)
        ytd_filtered = filter_combined(ytd, d_ids, funcs, f_leads, t_leads, months)
        ytd_filtered = apply_search(ytd_filtered, search)
        ytd_filtered = clean_dataframe_for_display(ytd_filtered, st.session_state.hide_cols)

        c1, c2, c3 = st.columns(3)
        c1.metric("YTD Rows (after filters)", len(ytd_filtered))
        c2.metric("Distinct Domains", ytd_filtered["Domain ID"].nunique() if "Domain ID" in ytd_filtered.columns else 0)
        c3.metric("Distinct Functions", ytd_filtered["Function"].nunique() if "Function" in ytd_filtered.columns else 0)

        st.subheader("YTD Aggregated BA Table")
        agg_options = [opt for opt in ["Name","Domain ID"] if opt in ytd_filtered.columns]
        agg_by = st.selectbox("Aggregate by", options=agg_options, index=0 if "Name" in agg_options else 0)
        ytd_ba_agg = ytd_aggregated_table(ytd_filtered, group_by=agg_by)
        st.caption(f"Showing {len(ytd_ba_agg)} aggregated rows")
        st.dataframe(ytd_ba_agg, height=420)
        if exceeds_excel_limits(ytd_ba_agg):
            st.caption("Note: Aggregated result is too wide for Excel; download provided as CSV.")
        st.download_button(
            "⬇️ Download aggregated (YTD BA Table)",
            make_excel_bytes_from_df(ytd_ba_agg, st.session_state.hide_cols),
            file_name=f"{BA_EXPORT_PREFIX}ytd_ba_aggregated_{agg_by.lower().replace(' ', '_')}.xlsx"
        )

        enable_altair_theme()
        with st.expander("🎨 Advanced Visualizations (YTD)", expanded=False):
            if ytd_filtered is None or ytd_filtered.empty:
                st.info("No YTD data under current filters for advanced visuals.")
            else:
                cset1, cset2, cset3, cset4 = st.columns([2,2,2,2])
                metric_options_ytd = get_numeric_metric_options(ytd_filtered)
                default_metric_list_ytd = metric_options_ytd if metric_options_ytd else ["Final Score_num"]
                default_index_ytd = default_metric_list_ytd.index("Final Score_num") if "Final Score_num" in default_metric_list_ytd else 0
                sel_metric_ytd = cset1.selectbox("Metric (numeric %)", options=default_metric_list_ytd, index=default_index_ytd)
                agg_method_ytd = cset2.radio("Aggregation", ["mean", "median"], index=0)
                dim_candidates_ytd = [c for c in ["Function","Team Lead","Function Lead","Domain ID"] if c in ytd_filtered.columns]
                dim_ytd = cset3.selectbox("Group by", options=dim_candidates_ytd if dim_candidates_ytd else ["Function"], index=0)
                palette_ytd = cset4.selectbox("Palette", options=list(PALETTES.keys()), index=list(PALETTES.keys()).index("Tableau10"))
                cN1, cN2, cN3 = st.columns([1,1,1])
                top_n_ytd = cN1.slider("Top N", min_value=5, max_value=50, value=15, step=5)
                ascending_ytd = cN2.checkbox("Show lowest first", value=False)
                show_labels_ytd = cN3.checkbox("Bar labels", value=True)

                agg_ytd = aggregate_df(ytd_filtered, dim=dim_ytd, metric=sel_metric_ytd, method=agg_method_ytd)
                agg_ytd = add_rank_and_topN(agg_ytd, dim=dim_ytd, metric=sel_metric_ytd, top_n=top_n_ytd, ascending=ascending_ytd)
                st.altair_chart(
                    bar_chart(agg_ytd, dim=dim_ytd, metric=sel_metric_ytd,
                              title=f"{agg_method_ytd.title()} {sel_metric_ytd.replace('_num',' (%)')} by {dim_ytd} (YTD)",
                              palette=palette_ytd, show_labels=show_labels_ytd),
                    use_container_width=True
                )

                cH1, cH2 = st.columns([1,1])
                bin_step_ytd = cH1.slider("Histogram bin step (percentage points)", 1, 20, 5, 1)
                ref_ytd = cH2.radio("Reference line", ["mean", "median"], index=0)
                st.altair_chart(
                    histogram(ytd_filtered, metric=sel_metric_ytd, bin_step=bin_step_ytd,
                              title=f"Distribution of {sel_metric_ytd.replace('_num',' (%)')} (YTD)",
                              reference=ref_ytd),
                    use_container_width=True
                )

                st.altair_chart(
                    boxplot(ytd_filtered, dim=dim_ytd, metric=sel_metric_ytd, title=f"Distribution by {dim_ytd} (YTD)",
                    ), use_container_width=True
                )

                if "Function" in ytd_filtered.columns and "Team Lead" in ytd_filtered.columns:
                    st.altair_chart(
                        heatmap(ytd_filtered, row_dim="Function", col_dim="Team Lead", metric=sel_metric_ytd,
                                title=f"Heatmap: {sel_metric_ytd.replace('_num',' (%)')} (Function x Team Lead) - YTD"),
                        use_container_width=True
                    )
                else:
                    ytd_norm = ytd_filtered.copy()
                    ytd_norm["Month_norm"] = _to_month_str_series(ytd_norm)
                    if "Function" in ytd_norm.columns and "Month_norm" in ytd_norm.columns:
                        st.altair_chart(
                            heatmap(ytd_norm, row_dim="Function", col_dim="Month_norm", metric=sel_metric_ytd,
                                    title=f"Heatmap: {sel_metric_ytd.replace('_num',' (%)')} (Function x Month) - YTD"),
                            use_container_width=True
                        )

        st.subheader("Filtered YTD Table (BA)")
        st.caption(f"Showing {len(ytd_filtered)} of {len(ytd)} rows")
        st.dataframe(ytd_filtered, height=480)
        if exceeds_excel_limits(ytd_filtered):
            st.caption("Note: Filtered result is too wide for Excel; download provided as CSV.")
        st.download_button(
            "⬇️ Download filtered (YTD)",
            make_excel_bytes_from_df(ytd_filtered, st.session_state.hide_cols),
            file_name=f"{BA_EXPORT_PREFIX}ytd_dashboard_filtered.xlsx"
        )


# -------------------------------------
# New Page: PE Scorecard (Monthly/YTD metrics)
# -------------------------------------
elif page == "PE Scorecard (Monthly/YTD metrics)":
    st.header("📊 Process Expert (PE) Scorecard (Monthly/YTD metrics)")
    mode = st.radio("View mode", ["Monthly", "YTD"], index=0, horizontal=True)

    if mode == "Monthly":
        latest_row, latest_id, latest_data = pe_get_latest_monthly_data()
        if latest_data is None or latest_data.empty:
            st.warning("No active PE file available.")
        else:
            latest_data = clean_dataframe_for_display(latest_data, st.session_state.hide_cols)
            latest_data = add_numeric_cached(latest_data)

            def render_shared_filters_pe(df, label="Filters (Monthly)"):
                month_str = _to_month_str_series(df)
                month_options = sorted([m for m in month_str.dropna().unique() if m and str(m).strip() != ""])
                with st.expander(label, expanded=True):
                    c1, c2, c3, c4, c5 = st.columns(5)
                    domain_options = sorted(df["Domain ID"].dropna().astype(str).unique()) if "Domain ID" in df.columns else []
                    func_options = sorted(df["Function"].dropna().astype(str).unique()) if "Function" in df.columns else []
                    flead_options = sorted(df["Function Lead"].dropna().astype(str).unique()) if "Function Lead" in df.columns else []
                    tlead_options = sorted(df["Team Lead"].dropna().astype(str).unique()) if "Team Lead" in df.columns else []
                    d_ids = c1.multiselect("Domain ID (select one to comment)", domain_options)
                    funcs = c2.multiselect("Function", func_options)
                    f_leads = c3.multiselect("Function Lead", flead_options)
                    t_leads = c4.multiselect("Team Lead", tlead_options)
                    months = c5.multiselect("Month (YYYY-MM)", month_options)

                    if d_ids and len(d_ids) == 1 and months and len(months) == 1:
                        sel_domain = str(d_ids[0]); sel_month = str(months[0])
                        st.markdown("**Monthly feedback/feedforward** (max 500 characters)")
                        fb = pe_load_feedback()
                        existing = fb[(fb["Domain ID"].astype(str) == sel_domain) & (fb["Month"].astype(str) == sel_month)]
                        existing_text = existing.iloc[0]["Feedback"] if not existing.empty else ""
                        comment = st.text_area(
                            "Enter feedback for the selected Domain ID & Month",
                            value=str(existing_text),
                            max_chars=MAX_FEEDBACK_CHARS,
                            height=120
                        )
                        password = st.text_input("Enter Team Lead password to confirm", type="password")
                        if st.button("Submit Comment"):
                            if password == FEEDBACK_PASSWORD:
                                sel_df = df[df["Domain ID"].astype(str) == sel_domain]
                                if not sel_df.empty:
                                    name_val = sel_df.iloc[0]["Name"] if "Name" in sel_df.columns else sel_domain
                                    tl_val = sel_df.iloc[0]["Team Lead"] if "Team Lead" in sel_df.columns else ""
                                    if comment and comment.strip():
                                        pe_upsert_feedback(sel_domain, str(name_val), sel_month, str(tl_val),
                                                          comment.strip(), st.session_state.username or "user")
                                        st.success("Feedback saved.")
                                    else:
                                        st.warning("Please enter a feedback comment before submitting.")
                                else:
                                    st.error("Selected Domain ID not found in current dataset.")
                            else:
                                st.error("Incorrect password. Feedback not saved.")
                    return d_ids, funcs, f_leads, t_leads, months, None

            d_ids, funcs, f_leads, t_leads, months, _ = render_shared_filters_pe(latest_data)
            filtered = filter_combined(latest_data, d_ids, funcs, f_leads, t_leads, months)
            filtered = clean_dataframe_for_display(filtered, st.session_state.hide_cols)

            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Active Month", latest_row["reporting_month"])
            c2.metric("Rows (after filters)", len(filtered))
            c3.metric("Uploader", latest_row["uploader"])
            c4.metric("Last Upload", latest_row["upload_dt"])

            st.subheader("PE Monthly Metrics")
            active_month = str(latest_row["reporting_month"])
            if months and len(months) == 1:
                active_month = str(months[0])

            mon_metrics = monthly_metrics_table_pe(filtered, report_month=active_month, group_by="Domain ID")
            st.caption(f"Showing {len(mon_metrics)} monthly rows (from filtered view)")
            st.dataframe(mon_metrics, height=420)
            if exceeds_excel_limits(mon_metrics):
                st.caption("Note: Monthly metrics are too wide for Excel; download provided as CSV.")
            st.download_button(
                "⬇️ Download Monthly Metrics",
                make_excel_bytes_from_df(mon_metrics, st.session_state.hide_cols),
                file_name=f"{PE_EXPORT_PREFIX}monthly_pe_metrics_{active_month}.xlsx"
            )

            # Simple charts (same as BA/Associates)
            if "Function" in filtered.columns and "Final Score_num" in filtered.columns:
                final_func = (
                    filtered.dropna(subset=["Final Score_num", "Function"])
                            .groupby("Function", as_index=False)["Final Score_num"].mean()
                            .rename(columns={"Final Score_num": "Avg Final Score (%)"})
                )
                sel_func = alt.selection_multi(fields=["Function"], bind="legend")
                chart_a = alt.Chart(final_func).mark_bar().encode(
                    x=alt.X("Avg Final Score (%):Q", title="Avg Final Score (%)"),
                    y=alt.Y("Function:N", sort="-x"),
                    color=alt.Color("Function:N", legend=alt.Legend(title="Click legend to filter")),
                    tooltip=["Function","Avg Final Score (%)"]
                ).add_selection(sel_func).properties(title="Avg Final Score by Function (Monthly)")
                st.altair_chart(chart_a, use_container_width=True)
            else:
                st.info("Final Score or Function column not found—'Avg Final Score by Function' chart skipped.")
            if "Final Score_num" in filtered.columns:
                st.altair_chart(
                    alt.Chart(filtered.dropna(subset=["Final Score_num"]))
                       .mark_bar()
                       .encode(
                           x=alt.X("Final Score_num:Q", bin=alt.Bin(step=5), title="Final Score (%)"),
                           y=alt.Y("count():Q", title="Count"),
                           tooltip=[alt.Tooltip("Final Score_num:Q", title="Final Score (%)"), alt.Tooltip("count():Q", title="Count")]
                       ).properties(title="Final Score Distribution (Monthly, 5% bins)"),
                    use_container_width=True
                )

            enable_altair_theme()
            # Advanced visuals identical in structure to other pages (reuse helpers)
            with st.expander("🎨 Advanced Visualizations (Monthly)", expanded=False):
                if filtered is None or filtered.empty:
                    st.info("No data under current filters for advanced visuals.")
                else:
                    cset1, cset2, cset3, cset4 = st.columns([2,2,2,2])
                    metric_options = get_numeric_metric_options(filtered)
                    default_metric_list = metric_options if metric_options else ["Final Score_num"]
                    default_index = default_metric_list.index("Final Score_num") if "Final Score_num" in default_metric_list else 0
                    sel_metric = cset1.selectbox("Metric (numeric %)", options=default_metric_list, index=default_index)
                    agg_method = cset2.radio("Aggregation", ["mean", "median"], index=0)
                    dim_candidates = [c for c in ["Function","Team Lead","Function Lead","Domain ID"] if c in filtered.columns]
                    dim = cset3.selectbox("Group by", options=dim_candidates if dim_candidates else ["Function"], index=0)
                    palette = cset4.selectbox("Palette", options=list(PALETTES.keys()), index=list(PALETTES.keys()).index("Tableau10"))
                    cN1, cN2, cN3 = st.columns([1,1,1])
                    top_n = cN1.slider("Top N", min_value=5, max_value=50, value=15, step=5)
                    ascending = cN2.checkbox("Show lowest first", value=False)
                    show_labels = cN3.checkbox("Bar labels", value=True)
                    agg_df = aggregate_df(filtered, dim=dim, metric=sel_metric, method=agg_method)
                    agg_df = add_rank_and_topN(agg_df, dim=dim, metric=sel_metric, top_n=top_n, ascending=ascending)
                    st.altair_chart(
                        bar_chart(agg_df, dim=dim, metric=sel_metric,
                                  title=f"{agg_method.title()} {sel_metric.replace('_num',' (%)')} by {dim} (Monthly)",
                                  palette=palette, show_labels=show_labels),
                        use_container_width=True
                    )
                    cH1, cH2 = st.columns([1,1])
                    bin_step = cH1.slider("Histogram bin step (percentage points)", 1, 20, 5, 1)
                    ref = cH2.radio("Reference line", ["mean", "median"], index=0)
                    st.altair_chart(
                        histogram(filtered, metric=sel_metric, bin_step=bin_step,
                                  title=f"Distribution of {sel_metric.replace('_num',' (%)')} (Monthly)",
                                  reference=ref),
                        use_container_width=True
                    )
                    st.altair_chart(
                        boxplot(filtered, dim=dim, metric=sel_metric, title=f"Distribution by {dim} (Monthly)"),
                        use_container_width=True
                    )
                    try:
                        history = pe_load_history_cached()
                        active_mask_trend = _coerce_active_bool(history.get("active", pd.Series([], dtype="object")))
                        active_ids = history[active_mask_trend][["id","reporting_month"]].rename(columns={"id":"Attachment ID"})
                        monthly_all = pe_load_combined_cached()
                        monthly_active = monthly_all.merge(active_ids, on="Attachment ID", how="inner")
                        monthly_active["reporting_month"] = monthly_active["reporting_month"].astype(str)
                        monthly_active = add_numeric_percent_columns(monthly_active)
                        st.altair_chart(
                            line_trend(monthly_active, metric=sel_metric, month_col="reporting_month",
                                       title=f"Trend by Month (active attachments): {sel_metric.replace('_num',' (%)')}"),
                            use_container_width=True
                        )
                    except Exception:
                        st.caption("Trend by month unavailable under current data.")

            st.subheader("Filtered Table (latest active PE file)")
            st.caption(f"Showing {len(filtered)} of {len(latest_data)} rows")
            st.dataframe(filtered if not filtered.empty else pd.DataFrame(), height=480)
            if exceeds_excel_limits(filtered):
                st.caption("Note: Filtered result is too wide for Excel; download provided as CSV.")
            st.download_button(
                "⬇️ Download filtered (Monthly)",
                make_excel_bytes_from_df(filtered, st.session_state.hide_cols),
                file_name=f"{PE_EXPORT_PREFIX}monthly_scorecard_filtered.xlsx"
            )

            if st.session_state.role == "admin":
                st.subheader("🛠 Admin — Edit Latest Active PE Data")
                st.caption("Edit values directly. Saving replaces the data for the latest active PE attachment in combined storage (not the original Excel file).")
                editable = st.data_editor(latest_data.copy(), num_rows="dynamic", use_container_width=True)
                if st.button("Save Admin Changes (PE)", type="primary"):
                    try:
                        edited = editable.copy()
                        edited = convert_percentage_columns(edited)
                        edited = add_numeric_percent_columns(edited)
                        edited["Attachment ID"] = latest_id
                        combined_df = pe_load_combined_cached()
                        combined_df = combined_df[combined_df["Attachment ID"] != latest_id]
                        combined_df = pd.concat([combined_df, edited], ignore_index=True)
                        pe_save_combined(combined_df)
                        pe_log_audit("Admin Save Edit (PE)", latest_id,
                                     latest_row["filename"] if "filename" in latest_row else "",
                                     st.session_state.username or "admin")
                        st.success("Admin changes saved to PE combined storage.")
                        invalidate_data_caches()  # ensure next UI run fetches fresh files
                    except Exception as e:
                        st.error(f"Failed to save admin changes: {e}")

    else:
        # YTD
        history = pe_load_history_cached()
        active_mask = _coerce_active_bool(history.get("active", pd.Series([], dtype="object")))
        active_ids = history[active_mask][["id","reporting_month"]].rename(columns={"id":"Attachment ID"})
        monthly_all = pe_load_combined_cached()
        ytd = monthly_all.merge(active_ids, on="Attachment ID", how="inner")
        if ytd.empty:
            st.warning("No PE YTD data.")
            st.stop()
        ytd = clean_dataframe_for_display(ytd, st.session_state.hide_cols)
        ytd = add_numeric_cached(ytd)

        def render_ytd_filters_pe(df, label="Filters (YTD)"):
            month_str = _to_month_str_series(df)
            month_options = sorted([m for m in month_str.dropna().unique() if m and str(m).strip() != ""]) if not month_str.empty else []
            def _fy_label(ym: str) -> str:
                try:
                    y, m = ym.split('-'); y = int(y); m = int(m)
                except Exception:
                    return ""
                if m >= 4: fy_start, fy_end = y, y+1
                else: fy_start, fy_end = y-1, y
                return f"FY{fy_start}-{str(fy_end)[-2:]}"
            with st.expander(label, expanded=True):
                c1, c2, c3, c4, c5 = st.columns(5)
                domain_options = sorted(df["Domain ID"].dropna().astype(str).unique()) if "Domain ID" in df.columns else []
                func_options = sorted(df["Function"].dropna().astype(str).unique()) if "Function" in df.columns else []
                flead_options = sorted(df["Function Lead"].dropna().astype(str).unique()) if "Function Lead" in df.columns else []
                tlead_options = sorted(df["Team Lead"].dropna().astype(str).unique()) if "Team Lead" in df.columns else []
                d_ids = c1.multiselect("Domain ID", domain_options)
                funcs = c2.multiselect("Function", func_options)
                f_leads = c3.multiselect("Function Lead", flead_options)
                t_leads = c4.multiselect("Team Lead", tlead_options)
                use_fy = c5.checkbox("Use Fiscal Year (Apr–Mar)", value=False)
                if use_fy and month_options:
                    fy_options = sorted({_fy_label(m) for m in month_options if _fy_label(m)})
                    sel_fy = c5.selectbox("Fiscal Year", options=fy_options, index=0 if fy_options else 0)
                    months = [m for m in month_options if _fy_label(m) == sel_fy]
                    st.caption(f"Months auto-selected for {sel_fy}: {', '.join(months)}")
                else:
                    months = c5.multiselect("Month (YYYY-MM)", month_options)
                search = st.text_input("🔎 Search across all columns (YTD)")
                return d_ids, funcs, f_leads, t_leads, months, search

        d_ids, funcs, f_leads, t_leads, months, search = render_ytd_filters_pe(ytd)
        ytd_filtered = filter_combined(ytd, d_ids, funcs, f_leads, t_leads, months)
        ytd_filtered = apply_search(ytd_filtered, search)
        ytd_filtered = clean_dataframe_for_display(ytd_filtered, st.session_state.hide_cols)

        c1, c2, c3 = st.columns(3)
        c1.metric("YTD Rows (after filters)", len(ytd_filtered))
        c2.metric("Distinct Domains", ytd_filtered["Domain ID"].nunique() if "Domain ID" in ytd_filtered.columns else 0)
        c3.metric("Distinct Functions", ytd_filtered["Function"].nunique() if "Function" in ytd_filtered.columns else 0)

        st.subheader("YTD Aggregated PE Table")
        agg_options = [opt for opt in ["Name","Domain ID"] if opt in ytd_filtered.columns]
        agg_by = st.selectbox("Aggregate by", options=agg_options, index=0 if "Name" in agg_options else 0)
        ytd_pe_agg = ytd_aggregated_table(ytd_filtered, group_by=agg_by)
        st.caption(f"Showing {len(ytd_pe_agg)} aggregated rows")
        st.dataframe(ytd_pe_agg, height=420)
        if exceeds_excel_limits(ytd_pe_agg):
            st.caption("Note: Aggregated result is too wide for Excel; download provided as CSV.")
        st.download_button(
            "⬇️ Download aggregated (YTD PE Table)",
            make_excel_bytes_from_df(ytd_pe_agg, st.session_state.hide_cols),
            file_name=f"{PE_EXPORT_PREFIX}ytd_pe_aggregated_{agg_by.lower().replace(' ', '_')}.xlsx"
        )

        enable_altair_theme()
        with st.expander("🎨 Advanced Visualizations (YTD)", expanded=False):
            if ytd_filtered is None or ytd_filtered.empty:
                st.info("No YTD data under current filters for advanced visuals.")
            else:
                cset1, cset2, cset3, cset4 = st.columns([2,2,2,2])
                metric_options_ytd = get_numeric_metric_options(ytd_filtered)
                default_metric_list_ytd = metric_options_ytd if metric_options_ytd else ["Final Score_num"]
                default_index_ytd = default_metric_list_ytd.index("Final Score_num") if "Final Score_num" in default_metric_list_ytd else 0
                sel_metric_ytd = cset1.selectbox("Metric (numeric %)", options=default_metric_list_ytd, index=default_index_ytd)
                agg_method_ytd = cset2.radio("Aggregation", ["mean", "median"], index=0)
                dim_candidates_ytd = [c for c in ["Function","Team Lead","Function Lead","Domain ID"] if c in ytd_filtered.columns]
                dim_ytd = cset3.selectbox("Group by", options=dim_candidates_ytd if dim_candidates_ytd else ["Function"], index=0)
                palette_ytd = cset4.selectbox("Palette", options=list(PALETTES.keys()), index=list(PALETTES.keys()).index("Tableau10"))
                cN1, cN2, cN3 = st.columns([1,1,1])
                top_n_ytd = cN1.slider("Top N", min_value=5, max_value=50, value=15, step=5)
                ascending_ytd = cN2.checkbox("Show lowest first", value=False)
                show_labels_ytd = cN3.checkbox("Bar labels", value=True)
                agg_ytd = aggregate_df(ytd_filtered, dim=dim_ytd, metric=sel_metric_ytd, method=agg_method_ytd)
                agg_ytd = add_rank_and_topN(agg_ytd, dim=dim_ytd, metric=sel_metric_ytd, top_n=top_n_ytd, ascending=ascending_ytd)
                st.altair_chart(
                    bar_chart(agg_ytd, dim=dim_ytd, metric=sel_metric_ytd,
                              title=f"{agg_method_ytd.title()} {sel_metric_ytd.replace('_num',' (%)')} by {dim_ytd} (YTD)",
                              palette=palette_ytd, show_labels=show_labels_ytd),
                    use_container_width=True
                )
                cH1, cH2 = st.columns([1,1])
                bin_step_ytd = cH1.slider("Histogram bin step (percentage points)", 1, 20, 5, 1)
                ref_ytd = cH2.radio("Reference line", ["mean", "median"], index=0)
                st.altair_chart(
                    histogram(ytd_filtered, metric=sel_metric_ytd, bin_step=bin_step_ytd,
                              title=f"Distribution of {sel_metric_ytd.replace('_num',' (%)')} (YTD)",
                              reference=ref_ytd),
                    use_container_width=True
                )
                st.altair_chart(
                    boxplot(ytd_filtered, dim=dim_ytd, metric=sel_metric_ytd, title=f"Distribution by {dim_ytd} (YTD)"),
                    use_container_width=True
                )
                if "Function" in ytd_filtered.columns and "Team Lead" in ytd_filtered.columns:
                    st.altair_chart(
                        heatmap(ytd_filtered, row_dim="Function", col_dim="Team Lead", metric=sel_metric_ytd,
                                title=f"Heatmap: {sel_metric_ytd.replace('_num',' (%)')} (Function x Team Lead) - YTD"),
                        use_container_width=True
                    )
                else:
                    ytd_norm = ytd_filtered.copy()
                    ytd_norm["Month_norm"] = _to_month_str_series(ytd_norm)
                    if "Function" in ytd_norm.columns and "Month_norm" in ytd_norm.columns:
                        st.altair_chart(
                            heatmap(ytd_norm, row_dim="Function", col_dim="Month_norm", metric=sel_metric_ytd,
                                    title=f"Heatmap: {sel_metric_ytd.replace('_num',' (%)')} (Function x Month) - YTD"),
                            use_container_width=True
                        )

        st.subheader("Filtered YTD Table (PE)")
        st.caption(f"Showing {len(ytd_filtered)} of {len(ytd)} rows")
        st.dataframe(ytd_filtered, height=480)
        if exceeds_excel_limits(ytd_filtered):
            st.caption("Note: Filtered result is too wide for Excel; download provided as CSV.")
        st.download_button(
            "⬇️ Download filtered (YTD)",
            make_excel_bytes_from_df(ytd_filtered, st.session_state.hide_cols),
            file_name=f"{PE_EXPORT_PREFIX}ytd_dashboard_filtered.xlsx"
        )


# -------------------------------------
# New Page: TL Scorecard (Monthly/YTD metrics)
# -------------------------------------
elif page == "TL Scorecard (Monthly/YTD metrics)":
    st.header("📊 Team Lead (TL) Scorecard (Monthly/YTD metrics)")
    mode = st.radio("View mode", ["Monthly", "YTD"], index=0, horizontal=True)

    if mode == "Monthly":
        latest_row, latest_id, latest_data = tl_get_latest_monthly_data()
        if latest_data is None or latest_data.empty:
            st.warning("No active TL file available.")
        else:
            latest_data = clean_dataframe_for_display(latest_data, st.session_state.hide_cols)
            latest_data = add_numeric_cached(latest_data)

            def render_shared_filters_tl(df, label="Filters (Monthly)"):
                month_str = _to_month_str_series(df)
                month_options = sorted([m for m in month_str.dropna().unique() if m and str(m).strip() != ""])
                with st.expander(label, expanded=True):
                    c1, c2, c3, c4, c5 = st.columns(5)
                    domain_options = sorted(df["Domain ID"].dropna().astype(str).unique()) if "Domain ID" in df.columns else []
                    func_options = sorted(df["Function"].dropna().astype(str).unique()) if "Function" in df.columns else []
                    flead_options = sorted(df["Function Lead"].dropna().astype(str).unique()) if "Function Lead" in df.columns else []
                    tlead_options = sorted(df["Team Lead"].dropna().astype(str).unique()) if "Team Lead" in df.columns else []
                    d_ids = c1.multiselect("Domain ID (select one to comment)", domain_options)
                    funcs = c2.multiselect("Function", func_options)
                    f_leads = c3.multiselect("Function Lead", flead_options)
                    t_leads = c4.multiselect("Team Lead", tlead_options)
                    months = c5.multiselect("Month (YYYY-MM)", month_options)

                    if d_ids and len(d_ids) == 1 and months and len(months) == 1:
                        sel_domain = str(d_ids[0]); sel_month = str(months[0])
                        st.markdown("**Monthly feedback/feedforward** (max 500 characters)")
                        fb = tl_load_feedback()
                        existing = fb[(fb["Domain ID"].astype(str) == sel_domain) & (fb["Month"].astype(str) == sel_month)]
                        existing_text = existing.iloc[0]["Feedback"] if not existing.empty else ""
                        comment = st.text_area(
                            "Enter feedback for the selected Domain ID & Month",
                            value=str(existing_text),
                            max_chars=MAX_FEEDBACK_CHARS,
                            height=120
                        )
                        password = st.text_input("Enter Team Lead password to confirm", type="password")
                        if st.button("Submit Comment"):
                            if password == FEEDBACK_PASSWORD:
                                sel_df = df[df["Domain ID"].astype(str) == sel_domain]
                                if not sel_df.empty:
                                    name_val = sel_df.iloc[0]["Name"] if "Name" in sel_df.columns else sel_domain
                                    tl_val = sel_df.iloc[0]["Team Lead"] if "Team Lead" in sel_df.columns else ""
                                    if comment and comment.strip():
                                        tl_upsert_feedback(sel_domain, str(name_val), sel_month, str(tl_val),
                                                          comment.strip(), st.session_state.username or "user")
                                        st.success("Feedback saved.")
                                    else:
                                        st.warning("Please enter a feedback comment before submitting.")
                                else:
                                    st.error("Selected Domain ID not found in current dataset.")
                            else:
                                st.error("Incorrect password. Feedback not saved.")
                    return d_ids, funcs, f_leads, t_leads, months, None

            d_ids, funcs, f_leads, t_leads, months, _ = render_shared_filters_tl(latest_data)
            filtered = filter_combined(latest_data, d_ids, funcs, f_leads, t_leads, months)
            filtered = clean_dataframe_for_display(filtered, st.session_state.hide_cols)

            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Active Month", latest_row["reporting_month"])
            c2.metric("Rows (after filters)", len(filtered))
            c3.metric("Uploader", latest_row["uploader"])
            c4.metric("Last Upload", latest_row["upload_dt"])

            st.subheader("TL Monthly Metrics")
            active_month = str(latest_row["reporting_month"])
            if months and len(months) == 1:
                active_month = str(months[0])

            mon_metrics = monthly_metrics_table_tl(filtered, report_month=active_month, group_by="Domain ID")
            st.caption(f"Showing {len(mon_metrics)} monthly rows (from filtered view)")
            st.dataframe(mon_metrics, height=420)
            if exceeds_excel_limits(mon_metrics):
                st.caption("Note: Monthly metrics are too wide for Excel; download provided as CSV.")
            st.download_button(
                "⬇️ Download Monthly Metrics",
                make_excel_bytes_from_df(mon_metrics, st.session_state.hide_cols),
                file_name=f"{TL_EXPORT_PREFIX}monthly_tl_metrics_{active_month}.xlsx"
            )

            # Simple charts (same as BA/Associates)
            if "Function" in filtered.columns and "Final Score_num" in filtered.columns:
                final_func = (
                    filtered.dropna(subset=["Final Score_num", "Function"])
                            .groupby("Function", as_index=False)["Final Score_num"].mean()
                            .rename(columns={"Final Score_num": "Avg Final Score (%)"})
                )
                sel_func = alt.selection_multi(fields=["Function"], bind="legend")
                chart_a = alt.Chart(final_func).mark_bar().encode(
                    x=alt.X("Avg Final Score (%):Q", title="Avg Final Score (%)"),
                    y=alt.Y("Function:N", sort="-x"),
                    color=alt.Color("Function:N", legend=alt.Legend(title="Click legend to filter")),
                    tooltip=["Function","Avg Final Score (%)"]
                ).add_selection(sel_func).properties(title="Avg Final Score by Function (Monthly)")
                st.altair_chart(chart_a, use_container_width=True)
            else:
                st.info("Final Score or Function column not found—'Avg Final Score by Function' chart skipped.")
            if "Final Score_num" in filtered.columns:
                st.altair_chart(
                    alt.Chart(filtered.dropna(subset=["Final Score_num"]))
                       .mark_bar()
                       .encode(
                           x=alt.X("Final Score_num:Q", bin=alt.Bin(step=5), title="Final Score (%)"),
                           y=alt.Y("count():Q", title="Count"),
                           tooltip=[alt.Tooltip("Final Score_num:Q", title="Final Score (%)"), alt.Tooltip("count():Q", title="Count")]
                       ).properties(title="Final Score Distribution (Monthly, 5% bins)"),
                    use_container_width=True
                )

            enable_altair_theme()
            # Advanced visuals identical in structure to other pages (reuse helpers)
            with st.expander("🎨 Advanced Visualizations (Monthly)", expanded=False):
                if filtered is None or filtered.empty:
                    st.info("No data under current filters for advanced visuals.")
                else:
                    cset1, cset2, cset3, cset4 = st.columns([2,2,2,2])
                    metric_options = get_numeric_metric_options(filtered)
                    default_metric_list = metric_options if metric_options else ["Final Score_num"]
                    default_index = default_metric_list.index("Final Score_num") if "Final Score_num" in default_metric_list else 0
                    sel_metric = cset1.selectbox("Metric (numeric %)", options=default_metric_list, index=default_index)
                    agg_method = cset2.radio("Aggregation", ["mean", "median"], index=0)
                    dim_candidates = [c for c in ["Function","Team Lead","Function Lead","Domain ID"] if c in filtered.columns]
                    dim = cset3.selectbox("Group by", options=dim_candidates if dim_candidates else ["Function"], index=0)
                    palette = cset4.selectbox("Palette", options=list(PALETTES.keys()), index=list(PALETTES.keys()).index("Tableau10"))
                    cN1, cN2, cN3 = st.columns([1,1,1])
                    top_n = cN1.slider("Top N", min_value=5, max_value=50, value=15, step=5)
                    ascending = cN2.checkbox("Show lowest first", value=False)
                    show_labels = cN3.checkbox("Bar labels", value=True)
                    agg_df = aggregate_df(filtered, dim=dim, metric=sel_metric, method=agg_method)
                    agg_df = add_rank_and_topN(agg_df, dim=dim, metric=sel_metric, top_n=top_n, ascending=ascending)
                    st.altair_chart(
                        bar_chart(agg_df, dim=dim, metric=sel_metric,
                                  title=f"{agg_method.title()} {sel_metric.replace('_num',' (%)')} by {dim} (Monthly)",
                                  palette=palette, show_labels=show_labels),
                        use_container_width=True
                    )
                    cH1, cH2 = st.columns([1,1])
                    bin_step = cH1.slider("Histogram bin step (percentage points)", 1, 20, 5, 1)
                    ref = cH2.radio("Reference line", ["mean", "median"], index=0)
                    st.altair_chart(
                        histogram(filtered, metric=sel_metric, bin_step=bin_step,
                                  title=f"Distribution of {sel_metric.replace('_num',' (%)')} (Monthly)",
                                  reference=ref),
                        use_container_width=True
                    )
                    st.altair_chart(
                        boxplot(filtered, dim=dim, metric=sel_metric, title=f"Distribution by {dim} (Monthly)"),
                        use_container_width=True
                    )
                    try:
                        history = tl_load_history_cached()
                        active_mask_trend = _coerce_active_bool(history.get("active", pd.Series([], dtype="object")))
                        active_ids = history[active_mask_trend][["id","reporting_month"]].rename(columns={"id":"Attachment ID"})
                        monthly_all = tl_load_combined_cached()
                        monthly_active = monthly_all.merge(active_ids, on="Attachment ID", how="inner")
                        monthly_active["reporting_month"] = monthly_active["reporting_month"].astype(str)
                        monthly_active = add_numeric_percent_columns(monthly_active)
                        st.altair_chart(
                            line_trend(monthly_active, metric=sel_metric, month_col="reporting_month",
                                       title=f"Trend by Month (active attachments): {sel_metric.replace('_num',' (%)')}"),
                            use_container_width=True
                        )
                    except Exception:
                        st.caption("Trend by month unavailable under current data.")

            st.subheader("Filtered Table (latest active TL file)")
            st.caption(f"Showing {len(filtered)} of {len(latest_data)} rows")
            st.dataframe(filtered if not filtered.empty else pd.DataFrame(), height=480)
            if exceeds_excel_limits(filtered):
                st.caption("Note: Filtered result is too wide for Excel; download provided as CSV.")
            st.download_button(
                "⬇️ Download filtered (Monthly)",
                make_excel_bytes_from_df(filtered, st.session_state.hide_cols),
                file_name=f"{TL_EXPORT_PREFIX}monthly_scorecard_filtered.xlsx"
            )

            if st.session_state.role == "admin":
                st.subheader("🛠 Admin — Edit Latest Active TL Data")
                st.caption("Edit values directly. Saving replaces the data for the latest active TL attachment in combined storage (not the original Excel file).")
                editable = st.data_editor(latest_data.copy(), num_rows="dynamic", use_container_width=True)
                if st.button("Save Admin Changes (TL)", type="primary"):
                    try:
                        edited = editable.copy()
                        edited = convert_percentage_columns(edited)
                        edited = add_numeric_percent_columns(edited)
                        edited["Attachment ID"] = latest_id
                        combined_df = tl_load_combined_cached()
                        combined_df = combined_df[combined_df["Attachment ID"] != latest_id]
                        combined_df = pd.concat([combined_df, edited], ignore_index=True)
                        tl_save_combined(combined_df)
                        tl_log_audit("Admin Save Edit (TL)", latest_id,
                                     latest_row["filename"] if "filename" in latest_row else "",
                                     st.session_state.username or "admin")
                        st.success("Admin changes saved to TL combined storage.")
                        invalidate_data_caches()  # ensure next UI run fetches fresh files
                    except Exception as e:
                        st.error(f"Failed to save admin changes: {e}")

    else:
        # YTD
        history = tl_load_history_cached()
        active_mask = _coerce_active_bool(history.get("active", pd.Series([], dtype="object")))
        active_ids = history[active_mask][["id","reporting_month"]].rename(columns={"id":"Attachment ID"})
        monthly_all = tl_load_combined_cached()
        ytd = monthly_all.merge(active_ids, on="Attachment ID", how="inner")
        if ytd.empty:
            st.warning("No TL YTD data.")
            st.stop()
        ytd = clean_dataframe_for_display(ytd, st.session_state.hide_cols)
        ytd = add_numeric_cached(ytd)

        def render_ytd_filters_tl(df, label="Filters (YTD)"):
            month_str = _to_month_str_series(df)
            month_options = sorted([m for m in month_str.dropna().unique() if m and str(m).strip() != ""]) if not month_str.empty else []
            def _fy_label(ym: str) -> str:
                try:
                    y, m = ym.split('-'); y = int(y); m = int(m)
                except Exception:
                    return ""
                if m >= 4: fy_start, fy_end = y, y+1
                else: fy_start, fy_end = y-1, y
                return f"FY{fy_start}-{str(fy_end)[-2:]}"
            with st.expander(label, expanded=True):
                c1, c2, c3, c4, c5 = st.columns(5)
                domain_options = sorted(df["Domain ID"].dropna().astype(str).unique()) if "Domain ID" in df.columns else []
                func_options = sorted(df["Function"].dropna().astype(str).unique()) if "Function" in df.columns else []
                flead_options = sorted(df["Function Lead"].dropna().astype(str).unique()) if "Function Lead" in df.columns else []
                tlead_options = sorted(df["Team Lead"].dropna().astype(str).unique()) if "Team Lead" in df.columns else []
                d_ids = c1.multiselect("Domain ID", domain_options)
                funcs = c2.multiselect("Function", func_options)
                f_leads = c3.multiselect("Function Lead", flead_options)
                t_leads = c4.multiselect("Team Lead", tlead_options)
                use_fy = c5.checkbox("Use Fiscal Year (Apr–Mar)", value=False)
                if use_fy and month_options:
                    fy_options = sorted({_fy_label(m) for m in month_options if _fy_label(m)})
                    sel_fy = c5.selectbox("Fiscal Year", options=fy_options, index=0 if fy_options else 0)
                    months = [m for m in month_options if _fy_label(m) == sel_fy]
                    st.caption(f"Months auto-selected for {sel_fy}: {', '.join(months)}")
                else:
                    months = c5.multiselect("Month (YYYY-MM)", month_options)
                search = st.text_input("🔎 Search across all columns (YTD)")
                return d_ids, funcs, f_leads, t_leads, months, search

        d_ids, funcs, f_leads, t_leads, months, search = render_ytd_filters_tl(ytd)
        ytd_filtered = filter_combined(ytd, d_ids, funcs, f_leads, t_leads, months)
        ytd_filtered = apply_search(ytd_filtered, search)
        ytd_filtered = clean_dataframe_for_display(ytd_filtered, st.session_state.hide_cols)

        c1, c2, c3 = st.columns(3)
        c1.metric("YTD Rows (after filters)", len(ytd_filtered))
        c2.metric("Distinct Domains", ytd_filtered["Domain ID"].nunique() if "Domain ID" in ytd_filtered.columns else 0)
        c3.metric("Distinct Functions", ytd_filtered["Function"].nunique() if "Function" in ytd_filtered.columns else 0)

        st.subheader("YTD Aggregated TL Table")
        agg_options = [opt for opt in ["Name","Domain ID"] if opt in ytd_filtered.columns]
        agg_by = st.selectbox("Aggregate by", options=agg_options, index=0 if "Name" in agg_options else 0)
        ytd_tl_agg = ytd_aggregated_table(ytd_filtered, group_by=agg_by)
        st.caption(f"Showing {len(ytd_tl_agg)} aggregated rows")
        st.dataframe(ytd_tl_agg, height=420)
        if exceeds_excel_limits(ytd_tl_agg):
            st.caption("Note: Aggregated result is too wide for Excel; download provided as CSV.")
        st.download_button(
            "⬇️ Download aggregated (YTD TL Table)",
            make_excel_bytes_from_df(ytd_tl_agg, st.session_state.hide_cols),
            file_name=f"{TL_EXPORT_PREFIX}ytd_tl_aggregated_{agg_by.lower().replace(' ', '_')}.xlsx"
        )

        enable_altair_theme()
        with st.expander("🎨 Advanced Visualizations (YTD)", expanded=False):
            if ytd_filtered is None or ytd_filtered.empty:
                st.info("No YTD data under current filters for advanced visuals.")
            else:
                cset1, cset2, cset3, cset4 = st.columns([2,2,2,2])
                metric_options_ytd = get_numeric_metric_options(ytd_filtered)
                default_metric_list_ytd = metric_options_ytd if metric_options_ytd else ["Final Score_num"]
                default_index_ytd = default_metric_list_ytd.index("Final Score_num") if "Final Score_num" in default_metric_list_ytd else 0
                sel_metric_ytd = cset1.selectbox("Metric (numeric %)", options=default_metric_list_ytd, index=default_index_ytd)
                agg_method_ytd = cset2.radio("Aggregation", ["mean", "median"], index=0)
                dim_candidates_ytd = [c for c in ["Function","Team Lead","Function Lead","Domain ID"] if c in ytd_filtered.columns]
                dim_ytd = cset3.selectbox("Group by", options=dim_candidates_ytd if dim_candidates_ytd else ["Function"], index=0)
                palette_ytd = cset4.selectbox("Palette", options=list(PALETTES.keys()), index=list(PALETTES.keys()).index("Tableau10"))
                cN1, cN2, cN3 = st.columns([1,1,1])
                top_n_ytd = cN1.slider("Top N", min_value=5, max_value=50, value=15, step=5)
                ascending_ytd = cN2.checkbox("Show lowest first", value=False)
                show_labels_ytd = cN3.checkbox("Bar labels", value=True)
                agg_ytd = aggregate_df(ytd_filtered, dim=dim_ytd, metric=sel_metric_ytd, method=agg_method_ytd)
                agg_ytd = add_rank_and_topN(agg_ytd, dim=dim_ytd, metric=sel_metric_ytd, top_n=top_n_ytd, ascending=ascending_ytd)
                st.altair_chart(
                    bar_chart(agg_ytd, dim=dim_ytd, metric=sel_metric_ytd,
                              title=f"{agg_method_ytd.title()} {sel_metric_ytd.replace('_num',' (%)')} by {dim_ytd} (YTD)",
                              palette=palette_ytd, show_labels=show_labels_ytd),
                    use_container_width=True
                )
                cH1, cH2 = st.columns([1,1])
                bin_step_ytd = cH1.slider("Histogram bin step (percentage points)", 1, 20, 5, 1)
                ref_ytd = cH2.radio("Reference line", ["mean", "median"], index=0)
                st.altair_chart(
                    histogram(ytd_filtered, metric=sel_metric_ytd, bin_step=bin_step_ytd,
                              title=f"Distribution of {sel_metric_ytd.replace('_num',' (%)')} (YTD)",
                              reference=ref_ytd),
                    use_container_width=True
                )
                st.altair_chart(
                    boxplot(ytd_filtered, dim=dim_ytd, metric=sel_metric_ytd, title=f"Distribution by {dim_ytd} (YTD)"),
                    use_container_width=True
                )
                if "Function" in ytd_filtered.columns and "Team Lead" in ytd_filtered.columns:
                    st.altair_chart(
                        heatmap(ytd_filtered, row_dim="Function", col_dim="Team Lead", metric=sel_metric_ytd,
                                title=f"Heatmap: {sel_metric_ytd.replace('_num',' (%)')} (Function x Team Lead) - YTD"),
                        use_container_width=True
                    )
                else:
                    ytd_norm = ytd_filtered.copy()
                    ytd_norm["Month_norm"] = _to_month_str_series(ytd_norm)
                    if "Function" in ytd_norm.columns and "Month_norm" in ytd_norm.columns:
                        st.altair_chart(
                            heatmap(ytd_norm, row_dim="Function", col_dim="Month_norm", metric=sel_metric_ytd,
                                    title=f"Heatmap: {sel_metric_ytd.replace('_num',' (%)')} (Function x Month) - YTD"),
                            use_container_width=True
                        )

        st.subheader("Filtered YTD Table (TL)")
        st.caption(f"Showing {len(ytd_filtered)} of {len(ytd)} rows")
        st.dataframe(ytd_filtered, height=480)
        if exceeds_excel_limits(ytd_filtered):
            st.caption("Note: Filtered result is too wide for Excel; download provided as CSV.")
        st.download_button(
            "⬇️ Download filtered (YTD)",
            make_excel_bytes_from_df(ytd_filtered, st.session_state.hide_cols),
            file_name=f"{TL_EXPORT_PREFIX}ytd_dashboard_filtered.xlsx"
        )


# -------------------------------------
# New Page: PL Scorecard (Monthly/YTD metrics)
# -------------------------------------
elif page == "PL Scorecard (Monthly/YTD metrics)":
    st.header("📊 Project Lead (PL) Scorecard (Monthly/YTD metrics)")
    mode = st.radio("View mode", ["Monthly", "YTD"], index=0, horizontal=True)

    if mode == "Monthly":
        latest_row, latest_id, latest_data = pl_get_latest_monthly_data()
        if latest_data is None or latest_data.empty:
            st.warning("No active PL file available.")
        else:
            latest_data = clean_dataframe_for_display(latest_data, st.session_state.hide_cols)
            latest_data = add_numeric_cached(latest_data)

            def render_shared_filters_pl(df, label="Filters (Monthly)"):
                month_str = _to_month_str_series(df)
                month_options = sorted([m for m in month_str.dropna().unique() if m and str(m).strip() != ""])
                with st.expander(label, expanded=True):
                    c1, c2, c3, c4, c5 = st.columns(5)
                    domain_options = sorted(df["Domain ID"].dropna().astype(str).unique()) if "Domain ID" in df.columns else []
                    func_options = sorted(df["Function"].dropna().astype(str).unique()) if "Function" in df.columns else []
                    flead_options = sorted(df["Function Lead"].dropna().astype(str).unique()) if "Function Lead" in df.columns else []
                    tlead_options = sorted(df["Team Lead"].dropna().astype(str).unique()) if "Team Lead" in df.columns else []
                    d_ids = c1.multiselect("Domain ID (select one to comment)", domain_options)
                    funcs = c2.multiselect("Function", func_options)
                    f_leads = c3.multiselect("Function Lead", flead_options)
                    t_leads = c4.multiselect("Team Lead", tlead_options)
                    months = c5.multiselect("Month (YYYY-MM)", month_options)

                    if d_ids and len(d_ids) == 1 and months and len(months) == 1:
                        sel_domain = str(d_ids[0]); sel_month = str(months[0])
                        st.markdown("**Monthly feedback/feedforward** (max 500 characters)")
                        fb = pl_load_feedback()
                        existing = fb[(fb["Domain ID"].astype(str) == sel_domain) & (fb["Month"].astype(str) == sel_month)]
                        existing_text = existing.iloc[0]["Feedback"] if not existing.empty else ""
                        comment = st.text_area(
                            "Enter feedback for the selected Domain ID & Month",
                            value=str(existing_text),
                            max_chars=MAX_FEEDBACK_CHARS,
                            height=120
                        )
                        password = st.text_input("Enter Team Lead password to confirm", type="password")
                        if st.button("Submit Comment"):
                            if password == FEEDBACK_PASSWORD:
                                sel_df = df[df["Domain ID"].astype(str) == sel_domain]
                                if not sel_df.empty:
                                    name_val = sel_df.iloc[0]["Name"] if "Name" in sel_df.columns else sel_domain
                                    tl_val = sel_df.iloc[0]["Team Lead"] if "Team Lead" in sel_df.columns else ""
                                    if comment and comment.strip():
                                        pl_upsert_feedback(sel_domain, str(name_val), sel_month, str(tl_val),
                                                          comment.strip(), st.session_state.username or "user")
                                        st.success("Feedback saved.")
                                    else:
                                        st.warning("Please enter a feedback comment before submitting.")
                                else:
                                    st.error("Selected Domain ID not found in current dataset.")
                            else:
                                st.error("Incorrect password. Feedback not saved.")
                    return d_ids, funcs, f_leads, t_leads, months, None

            d_ids, funcs, f_leads, t_leads, months, _ = render_shared_filters_pl(latest_data)
            filtered = filter_combined(latest_data, d_ids, funcs, f_leads, t_leads, months)
            filtered = clean_dataframe_for_display(filtered, st.session_state.hide_cols)

            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Active Month", latest_row["reporting_month"])
            c2.metric("Rows (after filters)", len(filtered))
            c3.metric("Uploader", latest_row["uploader"])
            c4.metric("Last Upload", latest_row["upload_dt"])

            st.subheader("PL Monthly Metrics")
            active_month = str(latest_row["reporting_month"])
            if months and len(months) == 1:
                active_month = str(months[0])

            mon_metrics = monthly_metrics_table_pl(filtered, report_month=active_month, group_by="Domain ID")
            st.caption(f"Showing {len(mon_metrics)} monthly rows (from filtered view)")
            st.dataframe(mon_metrics, height=420)
            if exceeds_excel_limits(mon_metrics):
                st.caption("Note: Monthly metrics are too wide for Excel; download provided as CSV.")
            st.download_button(
                "⬇️ Download Monthly Metrics",
                make_excel_bytes_from_df(mon_metrics, st.session_state.hide_cols),
                file_name=f"{PL_EXPORT_PREFIX}monthly_pl_metrics_{active_month}.xlsx"
            )

            # Simple charts (same as BA/Associates)
            if "Function" in filtered.columns and "Final Score_num" in filtered.columns:
                final_func = (
                    filtered.dropna(subset=["Final Score_num", "Function"])
                            .groupby("Function", as_index=False)["Final Score_num"].mean()
                            .rename(columns={"Final Score_num": "Avg Final Score (%)"})
                )
                sel_func = alt.selection_multi(fields=["Function"], bind="legend")
                chart_a = alt.Chart(final_func).mark_bar().encode(
                    x=alt.X("Avg Final Score (%):Q", title="Avg Final Score (%)"),
                    y=alt.Y("Function:N", sort="-x"),
                    color=alt.Color("Function:N", legend=alt.Legend(title="Click legend to filter")),
                    tooltip=["Function","Avg Final Score (%)"]
                ).add_selection(sel_func).properties(title="Avg Final Score by Function (Monthly)")
                st.altair_chart(chart_a, use_container_width=True)
            else:
                st.info("Final Score or Function column not found—'Avg Final Score by Function' chart skipped.")
            if "Final Score_num" in filtered.columns:
                st.altair_chart(
                    alt.Chart(filtered.dropna(subset=["Final Score_num"]))
                       .mark_bar()
                       .encode(
                           x=alt.X("Final Score_num:Q", bin=alt.Bin(step=5), title="Final Score (%)"),
                           y=alt.Y("count():Q", title="Count"),
                           tooltip=[alt.Tooltip("Final Score_num:Q", title="Final Score (%)"), alt.Tooltip("count():Q", title="Count")]
                       ).properties(title="Final Score Distribution (Monthly, 5% bins)"),
                    use_container_width=True
                )

            enable_altair_theme()
            # Advanced visuals identical in structure to other pages (reuse helpers)
            with st.expander("🎨 Advanced Visualizations (Monthly)", expanded=False):
                if filtered is None or filtered.empty:
                    st.info("No data under current filters for advanced visuals.")
                else:
                    cset1, cset2, cset3, cset4 = st.columns([2,2,2,2])
                    metric_options = get_numeric_metric_options(filtered)
                    default_metric_list = metric_options if metric_options else ["Final Score_num"]
                    default_index = default_metric_list.index("Final Score_num") if "Final Score_num" in default_metric_list else 0
                    sel_metric = cset1.selectbox("Metric (numeric %)", options=default_metric_list, index=default_index)
                    agg_method = cset2.radio("Aggregation", ["mean", "median"], index=0)
                    dim_candidates = [c for c in ["Function","Team Lead","Function Lead","Domain ID"] if c in filtered.columns]
                    dim = cset3.selectbox("Group by", options=dim_candidates if dim_candidates else ["Function"], index=0)
                    palette = cset4.selectbox("Palette", options=list(PALETTES.keys()), index=list(PALETTES.keys()).index("Tableau10"))
                    cN1, cN2, cN3 = st.columns([1,1,1])
                    top_n = cN1.slider("Top N", min_value=5, max_value=50, value=15, step=5)
                    ascending = cN2.checkbox("Show lowest first", value=False)
                    show_labels = cN3.checkbox("Bar labels", value=True)
                    agg_df = aggregate_df(filtered, dim=dim, metric=sel_metric, method=agg_method)
                    agg_df = add_rank_and_topN(agg_df, dim=dim, metric=sel_metric, top_n=top_n, ascending=ascending)
                    st.altair_chart(
                        bar_chart(agg_df, dim=dim, metric=sel_metric,
                                  title=f"{agg_method.title()} {sel_metric.replace('_num',' (%)')} by {dim} (Monthly)",
                                  palette=palette, show_labels=show_labels),
                        use_container_width=True
                    )
                    cH1, cH2 = st.columns([1,1])
                    bin_step = cH1.slider("Histogram bin step (percentage points)", 1, 20, 5, 1)
                    ref = cH2.radio("Reference line", ["mean", "median"], index=0)
                    st.altair_chart(
                        histogram(filtered, metric=sel_metric, bin_step=bin_step,
                                  title=f"Distribution of {sel_metric.replace('_num',' (%)')} (Monthly)",
                                  reference=ref),
                        use_container_width=True
                    )
                    st.altair_chart(
                        boxplot(filtered, dim=dim, metric=sel_metric, title=f"Distribution by {dim} (Monthly)"),
                        use_container_width=True
                    )
                    try:
                        history = pl_load_history_cached()
                        active_mask_trend = _coerce_active_bool(history.get("active", pd.Series([], dtype="object")))
                        active_ids = history[active_mask_trend][["id","reporting_month"]].rename(columns={"id":"Attachment ID"})
                        monthly_all = pl_load_combined_cached()
                        monthly_active = monthly_all.merge(active_ids, on="Attachment ID", how="inner")
                        monthly_active["reporting_month"] = monthly_active["reporting_month"].astype(str)
                        monthly_active = add_numeric_percent_columns(monthly_active)
                        st.altair_chart(
                            line_trend(monthly_active, metric=sel_metric, month_col="reporting_month",
                                       title=f"Trend by Month (active attachments): {sel_metric.replace('_num',' (%)')}"),
                            use_container_width=True
                        )
                    except Exception:
                        st.caption("Trend by month unavailable under current data.")

            st.subheader("Filtered Table (latest active PL file)")
            st.caption(f"Showing {len(filtered)} of {len(latest_data)} rows")
            st.dataframe(filtered if not filtered.empty else pd.DataFrame(), height=480)
            if exceeds_excel_limits(filtered):
                st.caption("Note: Filtered result is too wide for Excel; download provided as CSV.")
            st.download_button(
                "⬇️ Download filtered (Monthly)",
                make_excel_bytes_from_df(filtered, st.session_state.hide_cols),
                file_name=f"{PL_EXPORT_PREFIX}monthly_scorecard_filtered.xlsx"
            )

            if st.session_state.role == "admin":
                st.subheader("🛠 Admin — Edit Latest Active PL Data")
                st.caption("Edit values directly. Saving replaces the data for the latest active PL attachment in combined storage (not the original Excel file).")
                editable = st.data_editor(latest_data.copy(), num_rows="dynamic", use_container_width=True)
                if st.button("Save Admin Changes (PL)", type="primary"):
                    try:
                        edited = editable.copy()
                        edited = convert_percentage_columns(edited)
                        edited = add_numeric_percent_columns(edited)
                        edited["Attachment ID"] = latest_id
                        combined_df = pl_load_combined_cached()
                        combined_df = combined_df[combined_df["Attachment ID"] != latest_id]
                        combined_df = pd.concat([combined_df, edited], ignore_index=True)
                        pl_save_combined(combined_df)
                        pl_log_audit("Admin Save Edit (PL)", latest_id,
                                     latest_row["filename"] if "filename" in latest_row else "",
                                     st.session_state.username or "admin")
                        st.success("Admin changes saved to PL combined storage.")
                        invalidate_data_caches()  # ensure next UI run fetches fresh files
                    except Exception as e:
                        st.error(f"Failed to save admin changes: {e}")

    else:
        # YTD
        history = pl_load_history_cached()
        active_mask = _coerce_active_bool(history.get("active", pd.Series([], dtype="object")))
        active_ids = history[active_mask][["id","reporting_month"]].rename(columns={"id":"Attachment ID"})
        monthly_all = pl_load_combined_cached()
        ytd = monthly_all.merge(active_ids, on="Attachment ID", how="inner")
        if ytd.empty:
            st.warning("No PL YTD data.")
            st.stop()
        ytd = clean_dataframe_for_display(ytd, st.session_state.hide_cols)
        ytd = add_numeric_cached(ytd)

        def render_ytd_filters_pl(df, label="Filters (YTD)"):
            month_str = _to_month_str_series(df)
            month_options = sorted([m for m in month_str.dropna().unique() if m and str(m).strip() != ""]) if not month_str.empty else []
            def _fy_label(ym: str) -> str:
                try:
                    y, m = ym.split('-'); y = int(y); m = int(m)
                except Exception:
                    return ""
                if m >= 4: fy_start, fy_end = y, y+1
                else: fy_start, fy_end = y-1, y
                return f"FY{fy_start}-{str(fy_end)[-2:]}"
            with st.expander(label, expanded=True):
                c1, c2, c3, c4, c5 = st.columns(5)
                domain_options = sorted(df["Domain ID"].dropna().astype(str).unique()) if "Domain ID" in df.columns else []
                func_options = sorted(df["Function"].dropna().astype(str).unique()) if "Function" in df.columns else []
                flead_options = sorted(df["Function Lead"].dropna().astype(str).unique()) if "Function Lead" in df.columns else []
                tlead_options = sorted(df["Team Lead"].dropna().astype(str).unique()) if "Team Lead" in df.columns else []
                d_ids = c1.multiselect("Domain ID", domain_options)
                funcs = c2.multiselect("Function", func_options)
                f_leads = c3.multiselect("Function Lead", flead_options)
                t_leads = c4.multiselect("Team Lead", tlead_options)
                use_fy = c5.checkbox("Use Fiscal Year (Apr–Mar)", value=False)
                if use_fy and month_options:
                    fy_options = sorted({_fy_label(m) for m in month_options if _fy_label(m)})
                    sel_fy = c5.selectbox("Fiscal Year", options=fy_options, index=0 if fy_options else 0)
                    months = [m for m in month_options if _fy_label(m) == sel_fy]
                    st.caption(f"Months auto-selected for {sel_fy}: {', '.join(months)}")
                else:
                    months = c5.multiselect("Month (YYYY-MM)", month_options)
                search = st.text_input("🔎 Search across all columns (YTD)")
                return d_ids, funcs, f_leads, t_leads, months, search

        d_ids, funcs, f_leads, t_leads, months, search = render_ytd_filters_pl(ytd)
        ytd_filtered = filter_combined(ytd, d_ids, funcs, f_leads, t_leads, months)
        ytd_filtered = apply_search(ytd_filtered, search)
        ytd_filtered = clean_dataframe_for_display(ytd_filtered, st.session_state.hide_cols)

        c1, c2, c3 = st.columns(3)
        c1.metric("YTD Rows (after filters)", len(ytd_filtered))
        c2.metric("Distinct Domains", ytd_filtered["Domain ID"].nunique() if "Domain ID" in ytd_filtered.columns else 0)
        c3.metric("Distinct Functions", ytd_filtered["Function"].nunique() if "Function" in ytd_filtered.columns else 0)

        st.subheader("YTD Aggregated PL Table")
        agg_options = [opt for opt in ["Name","Domain ID"] if opt in ytd_filtered.columns]
        agg_by = st.selectbox("Aggregate by", options=agg_options, index=0 if "Name" in agg_options else 0)
        ytd_pl_agg = ytd_aggregated_table(ytd_filtered, group_by=agg_by)
        st.caption(f"Showing {len(ytd_pl_agg)} aggregated rows")
        st.dataframe(ytd_pl_agg, height=420)
        if exceeds_excel_limits(ytd_pl_agg):
            st.caption("Note: Aggregated result is too wide for Excel; download provided as CSV.")
        st.download_button(
            "⬇️ Download aggregated (YTD PL Table)",
            make_excel_bytes_from_df(ytd_pl_agg, st.session_state.hide_cols),
            file_name=f"{PL_EXPORT_PREFIX}ytd_pl_aggregated_{agg_by.lower().replace(' ', '_')}.xlsx"
        )

        enable_altair_theme()
        with st.expander("🎨 Advanced Visualizations (YTD)", expanded=False):
            if ytd_filtered is None or ytd_filtered.empty:
                st.info("No YTD data under current filters for advanced visuals.")
            else:
                cset1, cset2, cset3, cset4 = st.columns([2,2,2,2])
                metric_options_ytd = get_numeric_metric_options(ytd_filtered)
                default_metric_list_ytd = metric_options_ytd if metric_options_ytd else ["Final Score_num"]
                default_index_ytd = default_metric_list_ytd.index("Final Score_num") if "Final Score_num" in default_metric_list_ytd else 0
                sel_metric_ytd = cset1.selectbox("Metric (numeric %)", options=default_metric_list_ytd, index=default_index_ytd)
                agg_method_ytd = cset2.radio("Aggregation", ["mean", "median"], index=0)
                dim_candidates_ytd = [c for c in ["Function","Team Lead","Function Lead","Domain ID"] if c in ytd_filtered.columns]
                dim_ytd = cset3.selectbox("Group by", options=dim_candidates_ytd if dim_candidates_ytd else ["Function"], index=0)
                palette_ytd = cset4.selectbox("Palette", options=list(PALETTES.keys()), index=list(PALETTES.keys()).index("Tableau10"))
                cN1, cN2, cN3 = st.columns([1,1,1])
                top_n_ytd = cN1.slider("Top N", min_value=5, max_value=50, value=15, step=5)
                ascending_ytd = cN2.checkbox("Show lowest first", value=False)
                show_labels_ytd = cN3.checkbox("Bar labels", value=True)
                agg_ytd = aggregate_df(ytd_filtered, dim=dim_ytd, metric=sel_metric_ytd, method=agg_method_ytd)
                agg_ytd = add_rank_and_topN(agg_ytd, dim=dim_ytd, metric=sel_metric_ytd, top_n=top_n_ytd, ascending=ascending_ytd)
                st.altair_chart(
                    bar_chart(agg_ytd, dim=dim_ytd, metric=sel_metric_ytd,
                              title=f"{agg_method_ytd.title()} {sel_metric_ytd.replace('_num',' (%)')} by {dim_ytd} (YTD)",
                              palette=palette_ytd, show_labels=show_labels_ytd),
                    use_container_width=True
                )
                cH1, cH2 = st.columns([1,1])
                bin_step_ytd = cH1.slider("Histogram bin step (percentage points)", 1, 20, 5, 1)
                ref_ytd = cH2.radio("Reference line", ["mean", "median"], index=0)
                st.altair_chart(
                    histogram(ytd_filtered, metric=sel_metric_ytd, bin_step=bin_step_ytd,
                              title=f"Distribution of {sel_metric_ytd.replace('_num',' (%)')} (YTD)",
                              reference=ref_ytd),
                    use_container_width=True
                )
                st.altair_chart(
                    boxplot(ytd_filtered, dim=dim_ytd, metric=sel_metric_ytd, title=f"Distribution by {dim_ytd} (YTD)"),
                    use_container_width=True
                )
                if "Function" in ytd_filtered.columns and "Project Lead" in ytd_filtered.columns:
                    st.altair_chart(
                        heatmap(ytd_filtered, row_dim="Function", col_dim="Team Lead", metric=sel_metric_ytd,
                                title=f"Heatmap: {sel_metric_ytd.replace('_num',' (%)')} (Function x Team Lead) - YTD"),
                        use_container_width=True
                    )
                else:
                    ytd_norm = ytd_filtered.copy()
                    ytd_norm["Month_norm"] = _to_month_str_series(ytd_norm)
                    if "Function" in ytd_norm.columns and "Month_norm" in ytd_norm.columns:
                        st.altair_chart(
                            heatmap(ytd_norm, row_dim="Function", col_dim="Month_norm", metric=sel_metric_ytd,
                                    title=f"Heatmap: {sel_metric_ytd.replace('_num',' (%)')} (Function x Month) - YTD"),
                            use_container_width=True
                        )

        st.subheader("Filtered YTD Table (PL)")
        st.caption(f"Showing {len(ytd_filtered)} of {len(ytd)} rows")
        st.dataframe(ytd_filtered, height=480)
        if exceeds_excel_limits(ytd_filtered):
            st.caption("Note: Filtered result is too wide for Excel; download provided as CSV.")
        st.download_button(
            "⬇️ Download filtered (YTD)",
            make_excel_bytes_from_df(ytd_filtered, st.session_state.hide_cols),
            file_name=f"{PL_EXPORT_PREFIX}ytd_dashboard_filtered.xlsx"
        )




# -------------------------------------
# History
# -------------------------------------
elif page == "History":
    st.header("🗄️ Historical Table — All Attachments")
    h = load_history()
    if not h.empty:
        st.dataframe(h.sort_values("upload_dt", ascending=False), height=600)
    else:
        st.info("No history yet.")

# -------------------------------------
# Upload & Admin (with Manage Attachments)
# -------------------------------------
elif page == "Upload & Admin":
    if st.session_state.role != "admin":
        st.error("No permission.")
        st.stop()
    st.header("🛠️ Upload & Admin")

    file = st.file_uploader("Upload Excel (.xlsx with 'Data' sheet)", type=["xlsx"])
    if file:
        try:
            data_df = read_excel_bytes(file.getvalue())
            st.write("Data Preview:")
            st.dataframe(clean_dataframe_for_display(convert_percentage_columns(data_df).head(20), st.session_state.hide_cols))
        except Exception as e:
            st.error(f"Preview failed: {e}")

    detected_dataset = None
    if file:
        fname_lower = file.name.lower()
        if "business analyst" in fname_lower:
            detected_dataset = "BA"
            st.info("Detected dataset: **Business Analyst** (will be routed to BA storage).")
        elif "associate" in fname_lower:
            detected_dataset = "Associates"
            st.info("Detected dataset: **Associates** (will be routed to Associates storage).")
        elif "process expert" in fname_lower:  # NEW
            detected_dataset = "PE"
            st.info("Detected dataset: **Process Expert** (will be routed to PE storage).")
        elif "team lead" in fname_lower:  # NEW
            detected_dataset = "TL"
            st.info("Detected dataset: **Team Lead** (will be routed to TL storage).")
        elif "project lead" in fname_lower:  # NEW
            detected_dataset = "PL"
            st.info("Detected dataset: **Project Lead** (will be routed to PL storage).")
        else:
            st.warning("Filename does not include 'Associate', 'Business Analyst', 'Process Expert', 'Team Lead' or 'Project Lead'. Please include one of these terms for proper routing.")

    if st.button("Process Upload", disabled=(file is None or detected_dataset is None)):
        if detected_dataset == "BA":
            ok, msg, pd_preview = ba_process_upload(file.name, file.getvalue(), st.session_state.username)
        elif detected_dataset == "PE":  # NEW
            ok, msg, pd_preview = pe_process_upload(file.name, file.getvalue(), st.session_state.username)
        elif detected_dataset == "TL":  # NEW
            ok, msg, pd_preview = tl_process_upload(file.name, file.getvalue(), st.session_state.username)
        elif detected_dataset == "PL":  # NEW
            ok, msg, pd_preview = pl_process_upload(file.name, file.getvalue(), st.session_state.username)
        else:
            ok, msg, pd_preview = process_upload(file.name, file.getvalue(), st.session_state.username)
        if ok:
            st.success(msg)
            st.dataframe(clean_dataframe_for_display(pd_preview, st.session_state.hide_cols))
        else:
            st.error(msg)

    st.divider()
    st.subheader("Manage Attachments")
    tab1, tab2, tab3, tab4, tab5 = st.tabs(["Associates", "Business Analyst", "Process Expert", "Team Lead", "Project Lead"])

    with tab1:
        history_df = load_history()
        if history_df.empty:
            st.info("No Associates attachments yet.")
        else:
            c1, c2 = st.columns([2,1])
            selected_id = c1.selectbox("Select Attachment ID (Associates)", history_df["id"].tolist())
            action = c2.radio("Action", ["Mark Invalid & Cleanup", "Mark Valid (rebuild indexes)"], index=0)
            make_active = st.checkbox("Make Active Again (when marking Valid)", value=False)
            if st.button("Run Action (Associates)", use_container_width=True):
                if action == "Mark Invalid & Cleanup":
                    ok, msg = mark_invalid_and_cleanup(selected_id, st.session_state.username)
                    st.success(msg) if ok else st.error(msg)
                else:
                    ok, msg = mark_valid_and_rebuild(selected_id, make_active, st.session_state.username)
                    st.success(msg) if ok else st.error(msg)

    with tab2:
        ba_history_df = ba_load_history_cached()
        if ba_history_df.empty:
            st.info("No BA attachments yet.")
        else:
            c1, c2 = st.columns([2,1])
            ba_selected_id = c1.selectbox("Select Attachment ID (BA)", ba_history_df["id"].tolist())
            ba_action = c2.radio("Action (BA)", ["Mark Invalid & Cleanup", "Mark Valid (rebuild indexes)"], index=0)
            ba_make_active = st.checkbox("Make Active Again (BA)", value=False)
            if st.button("Run Action (BA)", use_container_width=True):
                if ba_action == "Mark Invalid & Cleanup":
                    ok, msg = ba_mark_invalid_and_cleanup(ba_selected_id, st.session_state.username)
                    st.success(msg) if ok else st.error(msg)
                else:
                    ok, msg = ba_mark_valid_and_rebuild(ba_selected_id, ba_make_active, st.session_state.username)
                    st.success(msg) if ok else st.error(msg)

    
    with tab3:
        pe_history_df = pe_load_history_cached()
        if pe_history_df.empty:
            st.info("No PE attachments yet.")
        else:
            c1, c2 = st.columns([2,1])
            pe_selected_id = c1.selectbox("Select Attachment ID (PE)", pe_history_df["id"].tolist())
            pe_action = c2.radio("Action (PE)", ["Mark Invalid & Cleanup", "Mark Valid (rebuild indexes)"], index=0)
            pe_make_active = st.checkbox("Make Active Again (PE)", value=False)
            if st.button("Run Action (PE)", use_container_width=True):
                if pe_action == "Mark Invalid & Cleanup":
                    ok, msg = pe_mark_invalid_and_cleanup(pe_selected_id, st.session_state.username)
                    st.success(msg) if ok else st.error(msg)
                else:
                    ok, msg = pe_mark_valid_and_rebuild(pe_selected_id, pe_make_active, st.session_state.username)
                    st.success(msg) if ok else st.error(msg)

    with tab4:
        tl_history_df = tl_load_history_cached()
        if tl_history_df.empty:
            st.info("No TL attachments yet.")
        else:
            c1, c2 = st.columns([2,1])
            tl_selected_id = c1.selectbox("Select Attachment ID (TL)", tl_history_df["id"].tolist())
            tl_action = c2.radio("Action (TL)", ["Mark Invalid & Cleanup", "Mark Valid (rebuild indexes)"], index=0)
            tl_make_active = st.checkbox("Make Active Again (TL)", value=False)
            if st.button("Run Action (TL)", use_container_width=True):
                if tl_action == "Mark Invalid & Cleanup":
                    ok, msg = tl_mark_invalid_and_cleanup(tl_selected_id, st.session_state.username)
                    st.success(msg) if ok else st.error(msg)
                else:
                    ok, msg = tl_mark_valid_and_rebuild(tl_selected_id, tl_make_active, st.session_state.username)
                    st.success(msg) if ok else st.error(msg)

    with tab5:
        pl_history_df = pl_load_history_cached()
        if pl_history_df.empty:
            st.info("No PL attachments yet.")
        else:
            c1, c2 = st.columns([2,1])
            pl_selected_id = c1.selectbox("Select Attachment ID (TL)", pl_history_df["id"].tolist())
            pl_action = c2.radio("Action (PL)", ["Mark Invalid & Cleanup", "Mark Valid (rebuild indexes)"], index=0)
            pl_make_active = st.checkbox("Make Active Again (PL)", value=False)
            if st.button("Run Action (PL)", use_container_width=True):
                if pl_action == "Mark Invalid & Cleanup":
                    ok, msg = pl_mark_invalid_and_cleanup(pl_selected_id, st.session_state.username)
                    st.success(msg) if ok else st.error(msg)
                else:
                    ok, msg = pl_mark_valid_and_rebuild(pl_selected_id, pl_make_active, st.session_state.username)
                    st.success(msg) if ok else st.error(msg)


st.divider()
st.subheader("Admin Notes")
st.markdown("""
- **Associates Scorecard (Monthly/YTD):** Single page to explore both modes; YTD filters include **Fiscal Year (Apr–Mar)** toggle.
- **BA Scorecard (Monthly/YTD):** Mirrors Associates visuals and filters with **independent storage/history/audit/feedback**.
- **PE Scorecard (Monthly/YTD):** Mirrors Associates/BA visuals and filters with **independent storage/history/audit/feedback**. YTD views are derived only from **active monthly attachments** (no `YTD` worksheet reads).
- **TL Scorecard (Monthly/YTD):** Mirrors Associates/BA/PE visuals and filters with **independent storage/history/audit/feedback**. YTD views are derived only from **active monthly attachments** (no `YTD` worksheet reads).
- **PL Scorecard (Monthly/YTD):** Mirrors Associates/BA/PE/TL visuals and filters with **independent storage/history/audit/feedback**. YTD views are derived only from **active monthly attachments** (no `YTD` worksheet reads).
- **Data Source:** All YTD views are derived from the **active monthly attachments** only (no `YTD` worksheet reads).
- **Monthly Metrics Table:** Built from the **filtered** latest active data; shows **Domain ID, Function, Function Lead, Team Lead, Designation, Name, Month, Final Score, Rank, Monthly feedback/feedforward, Feedback timestamp**.
- **YTD Aggregated Table:** Shows **Domain ID, Function, Function Lead, Team Lead, Designation, Name, Final Score, Rank** with **Final Score (mean)** across months and **dense rank**.
- **Filename routing:** Use "Associate" in the filename for Associates uploads; use "Business Analyst" for BA uploads.
- **Global toggle:** Hide/Show 'Unnamed' & fully empty columns across all pages & downloads.
- **Exports:** Filenames use dataset-specific prefixes (`associates_` / `ba_`). If tables exceed Excel size limits, downloads switch to **CSV**.
""")
